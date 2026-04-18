"""
JP BrokerFlow v2 API — multi-tenant endpoints.

Every endpoint requires a Supabase JWT and uses the user's stored
ClickUp token + other API keys from user_configs.

Mounted into the main FastAPI app as an APIRouter with prefix /api/v2.
"""

from __future__ import annotations

import json
import os
from typing import Any

import secrets
from datetime import datetime, timedelta, timezone

import httpx
from fastapi import APIRouter, Depends, HTTPException

from auth import (
    UserContext,
    get_current_user,
    upsert_user_config,
    supabase_configured,
    user_cu_get,
    user_cu_post,
    user_cu_put,
    user_cu_delete,
    SUPABASE_URL,
    SUPABASE_ANON_KEY,
)


router = APIRouter(prefix="/api/v2")


# =============================================================
# Masked key utility
# =============================================================

SENSITIVE_KEYS = {
    "clickup_token",
    "gmail_refresh_token",
    "smtp_pass",
    "twilio_auth_token",
    "anthropic_api_key",
}


def _mask(value: str | None) -> str | None:
    if not value:
        return None
    if len(value) <= 8:
        return "•" * len(value)
    return "•" * (len(value) - 4) + value[-4:]


def _mask_config(config: dict) -> dict:
    out = dict(config)
    for k in SENSITIVE_KEYS:
        if out.get(k):
            out[k] = _mask(out[k])
            out[f"{k}_set"] = True
        else:
            out[f"{k}_set"] = False
    return out


# =============================================================
# Config endpoints (Settings page)
# =============================================================

@router.get("/config")
async def get_config(user: UserContext = Depends(get_current_user)):
    """Return user's config with sensitive fields masked."""
    return _mask_config(user.config)


@router.put("/config")
async def update_config(
    payload: dict,
    user: UserContext = Depends(get_current_user),
):
    """Update any subset of user config fields.
    Empty string "" clears a field; omitted fields stay unchanged.
    Masked values (contain •) are ignored so Settings page re-save doesn't wipe keys.
    """
    allowed = {
        "clickup_token",
        "clickup_workspace_id",
        "clickup_space_id",
        "clickup_folder_id",
        "clickup_list_active_deals",
        "clickup_list_brokers",
        "clickup_list_followups",
        "clickup_list_templates",
        "clickup_list_touchpoints",
        "gmail_refresh_token",
        "gmail_email",
        "smtp_host",
        "smtp_port",
        "smtp_user",
        "smtp_pass",
        "smtp_from",
        "twilio_account_sid",
        "twilio_auth_token",
        "twilio_from_number",
        "anthropic_api_key",
        "wizard_step",
        "wizard_completed",
    }
    updates = {}
    for k, v in payload.items():
        if k not in allowed:
            continue
        if isinstance(v, str) and "•" in v:
            continue  # masked value, skip
        updates[k] = v

    if not updates:
        return _mask_config(user.config)

    row = await upsert_user_config(user.user_id, user.jwt, updates)
    return _mask_config(row)


# =============================================================
# Wizard: ClickUp
# =============================================================

@router.post("/wizard/clickup/test")
async def wizard_clickup_test(
    payload: dict,
    user: UserContext = Depends(get_current_user),
):
    """Verify a ClickUp token + return workspaces. Doesn't save yet."""
    token = (payload.get("token") or user.clickup_token or "").strip()
    if not token:
        raise HTTPException(400, "No token provided")

    headers = {"Authorization": token, "Content-Type": "application/json"}
    async with httpx.AsyncClient(timeout=15.0) as client:
        r = await client.get("https://api.clickup.com/api/v2/team", headers=headers)
        if r.status_code == 401:
            raise HTTPException(401, "ClickUp token invalid")
        if r.status_code != 200:
            raise HTTPException(r.status_code, f"ClickUp error: {r.text[:200]}")
        teams = r.json().get("teams", [])

    workspaces = [
        {"id": t["id"], "name": t.get("name", ""), "color": t.get("color", "")}
        for t in teams
    ]
    return {"ok": True, "workspaces": workspaces}


@router.post("/wizard/clickup/spaces")
async def wizard_clickup_spaces(
    payload: dict,
    user: UserContext = Depends(get_current_user),
):
    """List spaces in a workspace so user can pick where to put JP BrokerFlow lists."""
    token = (payload.get("token") or user.clickup_token or "").strip()
    workspace_id = payload.get("workspace_id", "").strip()
    if not token or not workspace_id:
        raise HTTPException(400, "token and workspace_id required")

    headers = {"Authorization": token, "Content-Type": "application/json"}
    async with httpx.AsyncClient(timeout=15.0) as client:
        r = await client.get(
            f"https://api.clickup.com/api/v2/team/{workspace_id}/space",
            headers=headers,
        )
        r.raise_for_status()
        spaces = r.json().get("spaces", [])
    return {"spaces": [{"id": s["id"], "name": s.get("name", "")} for s in spaces]}


@router.post("/wizard/clickup/setup")
async def wizard_clickup_setup(
    payload: dict,
    user: UserContext = Depends(get_current_user),
):
    """Create the JP BrokerFlow folder + 5 lists in the chosen space.
    Idempotent-ish: skips if folder named 'JP BrokerFlow' already exists.
    """
    token = (payload.get("token") or user.clickup_token or "").strip()
    workspace_id = payload.get("workspace_id", "").strip()
    space_id = payload.get("space_id", "").strip()
    if not all([token, workspace_id, space_id]):
        raise HTTPException(400, "token, workspace_id, space_id required")

    headers = {"Authorization": token, "Content-Type": "application/json"}
    folder_name = "JP BrokerFlow"
    list_specs = [
        ("clickup_list_active_deals", "Active Deals"),
        ("clickup_list_brokers", "Broker Directory"),
        ("clickup_list_followups", "Follow-ups Queue"),
        ("clickup_list_templates", "Message Templates"),
        ("clickup_list_touchpoints", "Touchpoints Log"),
    ]

    async with httpx.AsyncClient(timeout=30.0) as client:
        # Look for existing folder
        r = await client.get(
            f"https://api.clickup.com/api/v2/space/{space_id}/folder", headers=headers
        )
        r.raise_for_status()
        folders = r.json().get("folders", [])
        folder = next((f for f in folders if f.get("name") == folder_name), None)

        if not folder:
            r = await client.post(
                f"https://api.clickup.com/api/v2/space/{space_id}/folder",
                headers=headers,
                json={"name": folder_name},
            )
            r.raise_for_status()
            folder = r.json()

        folder_id = folder["id"]

        # Existing lists in folder
        r = await client.get(
            f"https://api.clickup.com/api/v2/folder/{folder_id}/list", headers=headers
        )
        r.raise_for_status()
        existing_lists = {l.get("name"): l for l in r.json().get("lists", [])}

        created_ids: dict[str, str] = {}
        for config_key, list_name in list_specs:
            if list_name in existing_lists:
                created_ids[config_key] = existing_lists[list_name]["id"]
                continue
            r = await client.post(
                f"https://api.clickup.com/api/v2/folder/{folder_id}/list",
                headers=headers,
                json={"name": list_name},
            )
            r.raise_for_status()
            created_ids[config_key] = r.json()["id"]

    updates = {
        "clickup_token": token,
        "clickup_workspace_id": workspace_id,
        "clickup_space_id": space_id,
        "clickup_folder_id": folder_id,
        **created_ids,
        "wizard_step": 2,
    }
    row = await upsert_user_config(user.user_id, user.jwt, updates)
    return {"ok": True, "folder_id": folder_id, "lists": created_ids, "config": _mask_config(row)}


# =============================================================
# Wizard: Anthropic
# =============================================================

@router.post("/wizard/anthropic/test")
async def wizard_anthropic_test(
    payload: dict,
    user: UserContext = Depends(get_current_user),
):
    """Verify an Anthropic API key by making a cheap messages call."""
    key = (payload.get("api_key") or user.config.get("anthropic_api_key") or "").strip()
    if not key:
        raise HTTPException(400, "No api_key provided")

    async with httpx.AsyncClient(timeout=15.0) as client:
        r = await client.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": key,
                "anthropic-version": "2023-06-01",
                "Content-Type": "application/json",
            },
            json={
                "model": "claude-haiku-4-5-20251001",
                "max_tokens": 10,
                "messages": [{"role": "user", "content": "ping"}],
            },
        )
        if r.status_code == 401:
            raise HTTPException(401, "Anthropic key invalid")
        if r.status_code >= 400:
            raise HTTPException(r.status_code, f"Anthropic error: {r.text[:200]}")

    await upsert_user_config(user.user_id, user.jwt, {"anthropic_api_key": key})
    return {"ok": True}


# =============================================================
# Wizard: Twilio
# =============================================================

@router.post("/wizard/twilio/test")
async def wizard_twilio_test(
    payload: dict,
    user: UserContext = Depends(get_current_user),
):
    """Verify Twilio creds by fetching the account record."""
    sid = (payload.get("account_sid") or user.config.get("twilio_account_sid") or "").strip()
    token = (payload.get("auth_token") or user.config.get("twilio_auth_token") or "").strip()
    from_number = (payload.get("from_number") or user.config.get("twilio_from_number") or "").strip()
    if not sid or not token:
        raise HTTPException(400, "account_sid and auth_token required")

    async with httpx.AsyncClient(timeout=15.0) as client:
        r = await client.get(
            f"https://api.twilio.com/2010-04-01/Accounts/{sid}.json",
            auth=(sid, token),
        )
        if r.status_code == 401:
            raise HTTPException(401, "Twilio creds invalid")
        if r.status_code >= 400:
            raise HTTPException(r.status_code, f"Twilio error: {r.text[:200]}")
        acct = r.json()

    updates = {
        "twilio_account_sid": sid,
        "twilio_auth_token": token,
    }
    if from_number:
        updates["twilio_from_number"] = from_number
    await upsert_user_config(user.user_id, user.jwt, updates)
    return {"ok": True, "account_name": acct.get("friendly_name", "")}


# =============================================================
# Wizard: SMTP (Gmail app-password / Outlook / custom)
# =============================================================

@router.post("/wizard/smtp/test")
async def wizard_smtp_test(
    payload: dict,
    user: UserContext = Depends(get_current_user),
):
    """Verify SMTP config by attempting auth (no message sent)."""
    import smtplib
    host = (payload.get("host") or user.config.get("smtp_host") or "").strip()
    port = int(payload.get("port") or user.config.get("smtp_port") or 587)
    smtp_user = (payload.get("user") or user.config.get("smtp_user") or "").strip()
    pw = (payload.get("password") or user.config.get("smtp_pass") or "").strip()
    smtp_from = (payload.get("from") or user.config.get("smtp_from") or smtp_user).strip()

    if not all([host, port, smtp_user, pw]):
        raise HTTPException(400, "host, port, user, password required")

    try:
        with smtplib.SMTP(host, port, timeout=15) as s:
            s.starttls()
            s.login(smtp_user, pw)
    except smtplib.SMTPAuthenticationError:
        raise HTTPException(401, "SMTP auth failed")
    except Exception as e:
        raise HTTPException(400, f"SMTP error: {e}")

    await upsert_user_config(
        user.user_id,
        user.jwt,
        {
            "smtp_host": host,
            "smtp_port": port,
            "smtp_user": smtp_user,
            "smtp_pass": pw,
            "smtp_from": smtp_from,
        },
    )
    return {"ok": True}


# =============================================================
# Wizard: complete
# =============================================================

@router.post("/wizard/complete")
async def wizard_complete(user: UserContext = Depends(get_current_user)):
    row = await upsert_user_config(
        user.user_id, user.jwt, {"wizard_completed": True, "wizard_step": 99}
    )
    return _mask_config(row)


# =============================================================
# v2 data endpoints (per-user ClickUp)
# =============================================================

# These are thin shims that delegate into broker_flow's mappers but use
# the user's token + lists. They're defined here to avoid circular imports
# by importing broker_flow lazily.

def _lazy_mappers():
    from broker_flow import task_to_deal, task_to_broker, task_to_followup
    return task_to_deal, task_to_broker, task_to_followup


BROKER_LIST_HINTS = ("broker", "contact", "people", "directory", "vendor", "agent")
SKIP_LIST_HINTS = ("sop", "template", "process", "playbook")


async def _fetch_space_lists(user: UserContext) -> list[dict]:
    """Return all lists in the user's configured ClickUp space, across all folders and folderless.

    Each item: {id, name, folder_id, folder_name, source}
    """
    space_id = (user.config.get("clickup_space_id") or "").strip()
    if not space_id:
        return []

    out: list[dict] = []
    # Folderless lists
    try:
        r = await user_cu_get(user, f"/space/{space_id}/list", params={"archived": "false"})
        for l in r.get("lists", []):
            out.append({
                "id": l.get("id"),
                "name": l.get("name", ""),
                "folder_id": None,
                "folder_name": None,
                "source": "folderless",
            })
    except Exception:
        pass

    # Folders + lists inside each folder
    try:
        fr = await user_cu_get(user, f"/space/{space_id}/folder", params={"archived": "false"})
        for folder in fr.get("folders", []):
            fid = folder.get("id")
            fname = folder.get("name", "")
            for l in folder.get("lists", []) or []:
                out.append({
                    "id": l.get("id"),
                    "name": l.get("name", ""),
                    "folder_id": fid,
                    "folder_name": fname,
                    "source": "folder",
                })
    except Exception:
        pass

    return [l for l in out if l.get("id")]


def _classify_list(name: str) -> str:
    n = (name or "").lower()
    if any(h in n for h in SKIP_LIST_HINTS):
        return "skip"
    if any(h in n for h in BROKER_LIST_HINTS):
        return "brokers"
    return "deals"


async def _fetch_tasks_for_list(user: UserContext, list_id: str) -> list[dict]:
    try:
        r = await user_cu_get(
            user,
            f"/list/{list_id}/task",
            params={"include_closed": "true", "subtasks": "true"},
        )
        return r.get("tasks", []) or []
    except Exception:
        return []


@router.get("/clickup/space-lists")
async def v2_space_lists(user: UserContext = Depends(get_current_user)):
    """Return every list in the connected space with classification hint.

    Lets the UI show users exactly which lists feed Deals vs Brokers.
    """
    lists = await _fetch_space_lists(user)
    for l in lists:
        l["classification"] = _classify_list(l.get("name", ""))
    return {"space_id": user.config.get("clickup_space_id"), "lists": lists}


# =============================================================
# WorkspaceContext — multi-tenant data routing
# =============================================================
#
# Every data endpoint (deals, brokers, touchpoints, links) resolves the
# "active workspace" from the X-Workspace-Id request header. The workspace
# owner's ClickUp/SMTP/Anthropic credentials are used to serve every
# member of that workspace — so when a user joins a workspace, they see
# the owner's deals, brokers, and history automatically (subject to role
# permissions).
#
# Fallback: if no header is sent, we use the caller's own default
# workspace (their oldest owned team) for backwards compatibility.

from dataclasses import dataclass as _dataclass
from fastapi import Header as _Header


@_dataclass
class WorkspaceContext:
    caller: UserContext                # the logged-in user
    owner: UserContext                 # workspace owner (with their config)
    workspace: dict                    # teams row
    membership: dict | None            # caller's team_members row (None if owner)
    perms: set                         # permissions caller has in this workspace
    workspace_id: str

    @property
    def is_owner(self) -> bool:
        return self.caller.user_id == self.workspace.get("owner_id")

    def has(self, perm: str) -> bool:
        return self.is_owner or perm in self.perms or "admin" in self.perms


async def _load_user_config_by_id(user_id: str) -> dict:
    """Service-role read of a user's config. Used to get workspace
    owner's ClickUp/SMTP/Anthropic keys to serve their members."""
    rows = await _svc_select(
        "user_configs",
        {"user_id": f"eq.{user_id}", "select": "*", "limit": "1"},
    )
    return rows[0] if rows else {}


async def _pick_default_workspace_id(user: UserContext) -> str | None:
    # 1. Prefer an owned workspace (oldest first)
    owned = await _svc_select(
        "teams",
        {
            "owner_id": f"eq.{user.user_id}",
            "select": "id",
            "order": "created_at.asc",
            "limit": "1",
        },
    )
    if owned:
        return owned[0]["id"]
    # 2. Else any workspace the caller is a member of
    mems = await _svc_select(
        "team_members",
        {"user_id": f"eq.{user.user_id}", "select": "team_id", "limit": "1"},
    )
    if mems:
        return mems[0]["team_id"]
    return None


async def get_workspace_context(
    x_workspace_id: str | None = _Header(default=None, alias="X-Workspace-Id"),
    user: UserContext = Depends(get_current_user),
) -> WorkspaceContext:
    """FastAPI dependency that resolves (caller, owner, workspace, perms)
    from the X-Workspace-Id header. Raises 403 if caller isn't a member."""
    ws_id = (x_workspace_id or "").strip() or None
    if not ws_id:
        ws_id = await _pick_default_workspace_id(user)
    if not ws_id:
        raise HTTPException(404, "No active workspace — create or join one first")

    team = await _get_team_or_404(ws_id)
    owner_id = team.get("owner_id")

    # Membership + perms
    membership: dict | None = None
    perms: set = set()
    if owner_id == user.user_id:
        perms = {"admin", "manage_members", "manage_roles", "manage_deals", "view_analytics"}
    else:
        m = await _get_membership(ws_id, user.user_id)
        if not m:
            raise HTTPException(403, "Not a member of this workspace")
        membership = m
        role = (m.get("team_roles") or {})
        role_perms = role.get("permissions") or {}
        perms = {k for k, v in role_perms.items() if v}

    # Build owner context (for ClickUp / SMTP / AI keys)
    if owner_id == user.user_id:
        owner_ctx = user
    else:
        owner_cfg = await _load_user_config_by_id(owner_id)
        # Synthetic UserContext — uses caller's JWT for any Supabase writes
        # (RLS sees the caller) but owner's config for external credentials.
        owner_ctx = UserContext(
            user_id=owner_id,
            email=owner_cfg.get("email") or "",
            is_anonymous=False,
            jwt=user.jwt,
            config=owner_cfg,
        )

    return WorkspaceContext(
        caller=user,
        owner=owner_ctx,
        workspace=team,
        membership=membership,
        perms=perms,
        workspace_id=ws_id,
    )


def require_ws_perm(perm: str):
    """Factory: returns a dependency that 403s if caller lacks perm."""
    async def _dep(
        wc: WorkspaceContext = Depends(get_workspace_context),
    ) -> WorkspaceContext:
        if not wc.has(perm):
            raise HTTPException(403, f"Missing permission: {perm}")
        return wc
    return _dep


async def _supabase_insert_ws(
    wc: WorkspaceContext, table: str, data: dict
) -> dict:
    """Insert a row stamped with workspace_id + user_id (caller).

    Uses the caller's JWT for RLS. workspace_id is the active workspace;
    user_id records who created the row (for audit)."""
    if not supabase_configured():
        raise HTTPException(503, "Supabase not configured")
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    headers = {
        "apikey": SUPABASE_ANON_KEY,
        "Authorization": f"Bearer {wc.caller.jwt}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }
    data = {
        "user_id": wc.caller.user_id,
        "workspace_id": wc.workspace_id,
        **data,
    }
    async with httpx.AsyncClient(timeout=15.0) as client:
        r = await client.post(url, headers=headers, json=data)
        if r.status_code not in (200, 201):
            raise HTTPException(r.status_code, f"Supabase insert failed: {r.text[:200]}")
        rows = r.json()
        return rows[0] if rows else {}


async def _supabase_rows_ws(
    wc: WorkspaceContext, table: str, params: dict[str, str] | None = None
) -> list[dict]:
    """Read rows from a workspace-scoped table. Uses caller's JWT so RLS
    applies. Caller must have already been verified as a workspace member
    (by get_workspace_context)."""
    if not supabase_configured():
        return []
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    headers = {
        "apikey": SUPABASE_ANON_KEY,
        "Authorization": f"Bearer {wc.caller.jwt}",
        "Accept": "application/json",
    }
    async with httpx.AsyncClient(timeout=15.0) as client:
        r = await client.get(url, headers=headers, params=params or {})
        if r.status_code != 200:
            return []
        return r.json()


@router.get("/deals")
async def v2_list_deals(wc: WorkspaceContext = Depends(get_workspace_context)):
    """Return all deals from the workspace-owner's connected ClickUp space.

    Sourcing rules (in order):
      1. If user configured a specific `clickup_list_active_deals` AND it has tasks, use only that.
      2. Else, union every list in the configured space whose name does NOT look like
         brokers/contacts/sops/templates. Each task is tagged with its source list name.
    """
    task_to_deal, _, _ = _lazy_mappers()
    owner = wc.owner

    # Try configured list first
    if owner.list_deals:
        try:
            r = await user_cu_get(
                owner,
                f"/list/{owner.list_deals}/task",
                params={"include_closed": "true", "subtasks": "true"},
            )
            tasks = r.get("tasks", []) or []
            if tasks:
                return {"deals": [task_to_deal(t) for t in tasks], "source": "configured_list"}
        except Exception:
            pass

    # Fall back to space-wide union
    lists = await _fetch_space_lists(owner)
    deals: list[dict] = []
    for l in lists:
        if _classify_list(l.get("name", "")) != "deals":
            continue
        list_id = l["id"]
        list_name = l.get("name", "")
        folder_name = l.get("folder_name")
        tasks = await _fetch_tasks_for_list(owner, list_id)
        for t in tasks:
            d = task_to_deal(t)
            d["source_list_id"] = list_id
            d["source_list_name"] = list_name
            d["source_folder_name"] = folder_name
            deals.append(d)
    return {"deals": deals, "source": "space_wide"}


@router.get("/brokers")
async def v2_list_brokers(wc: WorkspaceContext = Depends(get_workspace_context)):
    """Return all brokers from the workspace-owner's connected space.

    Sourcing rules (in order):
      1. If owner configured a specific `clickup_list_brokers` AND it has tasks, use only that.
      2. Else, union every list in the configured space whose name looks like brokers/contacts/people.
    """
    _, task_to_broker, _ = _lazy_mappers()
    owner = wc.owner

    if owner.list_brokers:
        try:
            r = await user_cu_get(
                owner, f"/list/{owner.list_brokers}/task", params={"include_closed": "true"}
            )
            tasks = r.get("tasks", []) or []
            if tasks:
                return {"brokers": [task_to_broker(t) for t in tasks], "source": "configured_list"}
        except Exception:
            pass

    lists = await _fetch_space_lists(owner)
    brokers: list[dict] = []
    for l in lists:
        if _classify_list(l.get("name", "")) != "brokers":
            continue
        list_id = l["id"]
        list_name = l.get("name", "")
        folder_name = l.get("folder_name")
        tasks = await _fetch_tasks_for_list(owner, list_id)
        for t in tasks:
            b = task_to_broker(t)
            b["source_list_id"] = list_id
            b["source_list_name"] = list_name
            b["source_folder_name"] = folder_name
            brokers.append(b)
    return {"brokers": brokers, "source": "space_wide"}


@router.get("/followups")
async def v2_list_followups(wc: WorkspaceContext = Depends(get_workspace_context)):
    owner = wc.owner
    if not owner.list_followups:
        return {"followups": [], "source": "not_configured"}
    _, _, task_to_followup = _lazy_mappers()
    try:
        resp = await user_cu_get(
            owner, f"/list/{owner.list_followups}/task", params={"include_closed": "true"}
        )
        return {"followups": [task_to_followup(t) for t in resp.get("tasks", [])], "source": "configured_list"}
    except Exception:
        return {"followups": [], "source": "error"}


# =============================================================
# Relationship view: deal + broker + recent touchpoints
# =============================================================

@router.get("/deals/{deal_id}/full")
async def v2_deal_full(deal_id: str, wc: WorkspaceContext = Depends(get_workspace_context)):
    """Deal + linked broker(s) + recent touchpoints (workspace-scoped)."""
    task_to_deal, task_to_broker, _ = _lazy_mappers()
    owner = wc.owner
    task = await user_cu_get(owner, f"/task/{deal_id}")
    deal = task_to_deal(task)

    brokers = []
    broker_ids: list[str] = []
    if deal.get("broker_id"):
        broker_ids.append(deal["broker_id"])

    # Also query deal_broker_links for many-to-many (scoped to workspace)
    link_rows = await _supabase_rows_ws(
        wc,
        "deal_broker_links",
        params={"workspace_id": f"eq.{wc.workspace_id}", "deal_id": f"eq.{deal_id}"},
    )
    for row in link_rows:
        bid = row.get("broker_id")
        if bid and bid not in broker_ids:
            broker_ids.append(bid)

    for bid in broker_ids:
        try:
            btask = await user_cu_get(owner, f"/task/{bid}")
            brokers.append(task_to_broker(btask))
        except Exception:
            continue

    touchpoints = await _supabase_rows_ws(
        wc,
        "touchpoints",
        params={
            "workspace_id": f"eq.{wc.workspace_id}",
            "deal_id": f"eq.{deal_id}",
            "order": "occurred_at.desc",
            "limit": "50",
        },
    )

    return {"deal": deal, "brokers": brokers, "touchpoints": touchpoints}


@router.get("/brokers/{broker_id}/full")
async def v2_broker_full(broker_id: str, wc: WorkspaceContext = Depends(get_workspace_context)):
    """Broker + associated deals + full activity timeline (workspace-scoped)."""
    task_to_deal, task_to_broker, _ = _lazy_mappers()
    owner = wc.owner
    task = await user_cu_get(owner, f"/task/{broker_id}")
    broker = task_to_broker(task)

    # All deals with this broker_id
    deals: list[dict] = []
    if owner.list_deals:
        resp = await user_cu_get(
            owner, f"/list/{owner.list_deals}/task", params={"include_closed": "true"}
        )
        for t in resp.get("tasks", []):
            d = task_to_deal(t)
            if d.get("broker_id") == broker_id:
                deals.append(d)

    # Plus from deal_broker_links (workspace-scoped)
    link_rows = await _supabase_rows_ws(
        wc,
        "deal_broker_links",
        params={"workspace_id": f"eq.{wc.workspace_id}", "broker_id": f"eq.{broker_id}"},
    )
    linked_deal_ids = {r["deal_id"] for r in link_rows}
    for did in linked_deal_ids:
        if any(d["id"] == did for d in deals):
            continue
        try:
            t = await user_cu_get(owner, f"/task/{did}")
            deals.append(task_to_deal(t))
        except Exception:
            continue

    touchpoints = await _supabase_rows_ws(
        wc,
        "touchpoints",
        params={
            "workspace_id": f"eq.{wc.workspace_id}",
            "broker_id": f"eq.{broker_id}",
            "order": "occurred_at.desc",
            "limit": "200",
        },
    )

    last_contact = touchpoints[0]["occurred_at"] if touchpoints else None
    return {
        "broker": broker,
        "deals": deals,
        "touchpoints": touchpoints,
        "last_contact": last_contact,
        "touchpoint_count": len(touchpoints),
    }


# =============================================================
# Touchpoints — log + query
# =============================================================

@router.post("/touchpoints")
async def v2_log_touchpoint(
    payload: dict,
    wc: WorkspaceContext = Depends(get_workspace_context),
):
    """Log a manual or auto touchpoint (workspace-scoped).
    payload: { broker_id, broker_name?, broker_email?, broker_phone?,
               deal_id?, deal_name?,
               channel: email|sms|call|note,
               direction: outbound|inbound|note,
               subject?, body?, duration_seconds?,
               source?: manual|dashboard|gmail_scan|twilio_webhook|smtp,
               occurred_at?: iso timestamp }
    """
    if not wc.has("manage_deals"):
        raise HTTPException(403, "Missing permission: manage_deals")

    required = ["broker_id", "channel", "direction"]
    for k in required:
        if not payload.get(k):
            raise HTTPException(400, f"Missing field: {k}")

    # Stamps workspace_id + user_id (caller) automatically
    row = await _supabase_insert_ws(wc, "touchpoints", payload)
    return row


@router.get("/touchpoints")
async def v2_query_touchpoints(
    broker_id: str | None = None,
    deal_id: str | None = None,
    limit: int = 100,
    wc: WorkspaceContext = Depends(get_workspace_context),
):
    params: dict[str, str] = {
        "workspace_id": f"eq.{wc.workspace_id}",
        "order": "occurred_at.desc",
        "limit": str(limit),
    }
    if broker_id:
        params["broker_id"] = f"eq.{broker_id}"
    if deal_id:
        params["deal_id"] = f"eq.{deal_id}"
    rows = await _supabase_rows_ws(wc, "touchpoints", params=params)
    return {"touchpoints": rows}


@router.delete("/touchpoints/{touchpoint_id}")
async def v2_delete_touchpoint(
    touchpoint_id: str,
    wc: WorkspaceContext = Depends(get_workspace_context),
):
    if not wc.has("manage_deals"):
        raise HTTPException(403, "Missing permission: manage_deals")
    await _supabase_delete(
        wc.caller,
        "touchpoints",
        params={"id": f"eq.{touchpoint_id}", "workspace_id": f"eq.{wc.workspace_id}"},
    )
    return {"ok": True}


# =============================================================
# AI drafting (Claude)
# =============================================================

@router.post("/draft/email")
async def v2_draft_email(
    payload: dict,
    wc: WorkspaceContext = Depends(get_workspace_context),
):
    """Generate a warm, JPIG-voiced doc-request / reminder email.
    payload: { broker_id, deal_id?, cadence_day?, purpose?, extra_context? }
    """
    owner = wc.owner
    key = owner.config.get("anthropic_api_key")
    if not key:
        raise HTTPException(400, "Anthropic key not set. Add it in Settings.")

    broker_id = payload.get("broker_id")
    if not broker_id:
        raise HTTPException(400, "broker_id required")

    _, task_to_broker, _ = _lazy_mappers()
    broker_task = await user_cu_get(owner, f"/task/{broker_id}")
    broker = task_to_broker(broker_task)

    deal_ctx = ""
    if payload.get("deal_id"):
        task_to_deal, _, _ = _lazy_mappers()
        try:
            deal_task = await user_cu_get(owner, f"/task/{payload['deal_id']}")
            deal = task_to_deal(deal_task)
            deal_ctx = (
                f"Deal: {deal.get('name','')}\n"
                f"Asset: {deal.get('asset_class','')}\n"
                f"Location: {deal.get('city','')}, {deal.get('state','')}\n"
                f"Ask: {deal.get('ask_price','')}\n"
                f"Stage: {deal.get('status','')}\n"
                f"Docs outstanding: {deal.get('docs_outstanding', [])}\n"
            )
        except Exception:
            pass

    # Recent conversation context (workspace-scoped)
    recent_tps = await _supabase_rows_ws(
        wc,
        "touchpoints",
        params={
            "workspace_id": f"eq.{wc.workspace_id}",
            "broker_id": f"eq.{broker_id}",
            "order": "occurred_at.desc",
            "limit": "5",
        },
    )
    history = "\n".join(
        f"- {t['occurred_at'][:10]} · {t['channel']} · {t['direction']}: {(t.get('subject') or t.get('body') or '')[:120]}"
        for t in recent_tps
    )

    purpose = payload.get("purpose", "doc request follow-up")
    cadence = payload.get("cadence_day", "")
    extra = payload.get("extra_context", "")

    system = (
        "You are writing a warm, professional email from Parth Patel at JP Investment "
        "Group (JPIG), an acquisitions shop. Voice: warm, relationship-first, "
        "low-pressure, brief. Never pushy. Prefer short paragraphs. End with a clear "
        "ask or next step. No marketing clichés. No 'I hope this email finds you well'. "
        "Sign 'Parth'."
    )
    user_msg = f"""Write an email to this broker.

BROKER
Name: {broker.get('name','')}
Firm: {broker.get('firm','')}
Relationship: {broker.get('relationship_strength','')}
Email: {broker.get('email','')}

DEAL CONTEXT
{deal_ctx or '(no specific deal)'}

PURPOSE: {purpose}
CADENCE DAY: {cadence or '(ad-hoc)'}
EXTRA: {extra}

RECENT TOUCHPOINTS
{history or '(none)'}

Return JSON: {{"subject": "...", "body": "..."}}"""

    async with httpx.AsyncClient(timeout=30.0) as client:
        r = await client.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": key,
                "anthropic-version": "2023-06-01",
                "Content-Type": "application/json",
            },
            json={
                "model": "claude-sonnet-4-5",
                "max_tokens": 800,
                "system": system,
                "messages": [{"role": "user", "content": user_msg}],
            },
        )
        if r.status_code >= 400:
            raise HTTPException(r.status_code, f"Anthropic error: {r.text[:200]}")
        data = r.json()
        text = data["content"][0]["text"]

    # Try to parse JSON response
    try:
        # Claude may wrap in ```json
        cleaned = text.strip()
        if cleaned.startswith("```"):
            cleaned = cleaned.split("```")[1]
            if cleaned.startswith("json"):
                cleaned = cleaned[4:]
            cleaned = cleaned.strip()
        parsed = json.loads(cleaned)
        subject = parsed.get("subject", "")
        body = parsed.get("body", "")
    except Exception:
        # Fallback: heuristic split
        lines = text.split("\n", 1)
        subject = lines[0].replace("Subject:", "").strip() if lines else ""
        body = lines[1].strip() if len(lines) > 1 else text

    return {"subject": subject, "body": body, "to": broker.get("email", "")}


@router.post("/draft/sms")
async def v2_draft_sms(
    payload: dict,
    wc: WorkspaceContext = Depends(get_workspace_context),
):
    """Generate a short SMS. payload: { broker_id, deal_id?, purpose? }"""
    owner = wc.owner
    key = owner.config.get("anthropic_api_key")
    if not key:
        raise HTTPException(400, "Anthropic key not set.")

    broker_id = payload.get("broker_id")
    if not broker_id:
        raise HTTPException(400, "broker_id required")

    _, task_to_broker, _ = _lazy_mappers()
    broker_task = await user_cu_get(owner, f"/task/{broker_id}")
    broker = task_to_broker(broker_task)

    purpose = payload.get("purpose", "quick check-in")

    system = (
        "Write a single SMS text message, max 160 chars, warm, casual, first-name "
        "basis. From Parth at JPIG. No links, no signature, no emoji unless natural."
    )
    user_msg = f"To: {broker.get('name','')} at {broker.get('firm','')}. Purpose: {purpose}."

    async with httpx.AsyncClient(timeout=20.0) as client:
        r = await client.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": key,
                "anthropic-version": "2023-06-01",
                "Content-Type": "application/json",
            },
            json={
                "model": "claude-haiku-4-5-20251001",
                "max_tokens": 200,
                "system": system,
                "messages": [{"role": "user", "content": user_msg}],
            },
        )
        r.raise_for_status()
        text = r.json()["content"][0]["text"].strip()

    return {"body": text, "to": broker.get("phone", "")}


# =============================================================
# Send: Email (SMTP) + SMS (Twilio)
# =============================================================

@router.post("/send/email")
async def v2_send_email(
    payload: dict,
    wc: WorkspaceContext = Depends(get_workspace_context),
):
    """Send an email via workspace owner's SMTP config. Logs a workspace-scoped touchpoint.
    payload: { to, subject, body, broker_id?, deal_id? }
    """
    if not wc.has("manage_deals"):
        raise HTTPException(403, "Missing permission: manage_deals")

    import smtplib
    from email.mime.text import MIMEText
    from email.utils import formatdate, make_msgid

    to = payload.get("to", "").strip()
    subject = payload.get("subject", "").strip()
    body = payload.get("body", "").strip()
    if not all([to, subject, body]):
        raise HTTPException(400, "to, subject, body required")

    owner = wc.owner
    host = owner.config.get("smtp_host")
    port = owner.config.get("smtp_port") or 587
    smtp_user = owner.config.get("smtp_user")
    pw = owner.config.get("smtp_pass")
    smtp_from = owner.config.get("smtp_from") or smtp_user
    if not all([host, smtp_user, pw]):
        raise HTTPException(400, "SMTP not configured. Run wizard.")

    msg = MIMEText(body, "plain", "utf-8")
    msg["Subject"] = subject
    msg["From"] = smtp_from
    msg["To"] = to
    msg["Date"] = formatdate(localtime=True)
    msg["Message-ID"] = make_msgid()

    try:
        with smtplib.SMTP(host, int(port), timeout=30) as s:
            s.starttls()
            s.login(smtp_user, pw)
            s.sendmail(smtp_from, [to], msg.as_string())
    except Exception as e:
        raise HTTPException(502, f"SMTP send failed: {e}")

    # Log touchpoint (workspace-scoped; user_id + workspace_id stamped by helper)
    tp_data = {
        "channel": "email",
        "direction": "outbound",
        "source": "smtp",
        "subject": subject,
        "body": body,
        "external_id": msg["Message-ID"],
    }
    if payload.get("broker_id"):
        tp_data["broker_id"] = payload["broker_id"]
    else:
        tp_data["broker_id"] = f"email:{to}"
    if payload.get("deal_id"):
        tp_data["deal_id"] = payload["deal_id"]
    tp_data["broker_email"] = to

    touchpoint = await _supabase_insert_ws(wc, "touchpoints", tp_data)
    return {"ok": True, "touchpoint": touchpoint}


@router.post("/send/sms")
async def v2_send_sms(
    payload: dict,
    wc: WorkspaceContext = Depends(get_workspace_context),
):
    """Send SMS via workspace owner's Twilio. Logs a workspace-scoped touchpoint.
    payload: { to, body, broker_id?, deal_id? }
    """
    if not wc.has("manage_deals"):
        raise HTTPException(403, "Missing permission: manage_deals")

    to = payload.get("to", "").strip()
    body = payload.get("body", "").strip()
    if not all([to, body]):
        raise HTTPException(400, "to, body required")

    owner = wc.owner
    sid = owner.config.get("twilio_account_sid")
    token = owner.config.get("twilio_auth_token")
    from_number = owner.config.get("twilio_from_number")
    if not all([sid, token, from_number]):
        raise HTTPException(400, "Twilio not configured.")

    async with httpx.AsyncClient(timeout=20.0) as client:
        r = await client.post(
            f"https://api.twilio.com/2010-04-01/Accounts/{sid}/Messages.json",
            auth=(sid, token),
            data={"From": from_number, "To": to, "Body": body},
        )
        if r.status_code >= 400:
            raise HTTPException(r.status_code, f"Twilio error: {r.text[:200]}")
        msg = r.json()

    tp_data = {
        "channel": "sms",
        "direction": "outbound",
        "source": "twilio_webhook",
        "body": body,
        "external_id": msg.get("sid"),
        "broker_phone": to,
    }
    if payload.get("broker_id"):
        tp_data["broker_id"] = payload["broker_id"]
    else:
        tp_data["broker_id"] = f"sms:{to}"
    if payload.get("deal_id"):
        tp_data["deal_id"] = payload["deal_id"]

    touchpoint = await _supabase_insert_ws(wc, "touchpoints", tp_data)
    return {"ok": True, "sid": msg.get("sid"), "touchpoint": touchpoint}


# =============================================================
# Supabase REST helpers (RLS-enforced via user JWT)
# =============================================================

async def _supabase_rows(
    user: UserContext,
    table: str,
    params: dict[str, str] | None = None,
) -> list[dict]:
    if not supabase_configured():
        return []
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    headers = {
        "apikey": SUPABASE_ANON_KEY,
        "Authorization": f"Bearer {user.jwt}",
        "Accept": "application/json",
    }
    async with httpx.AsyncClient(timeout=15.0) as client:
        r = await client.get(url, headers=headers, params=params or {})
        if r.status_code != 200:
            return []
        return r.json()


async def _supabase_insert(
    user: UserContext, table: str, data: dict
) -> dict:
    if not supabase_configured():
        raise HTTPException(503, "Supabase not configured")
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    headers = {
        "apikey": SUPABASE_ANON_KEY,
        "Authorization": f"Bearer {user.jwt}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }
    data = {"user_id": user.user_id, **data}
    async with httpx.AsyncClient(timeout=15.0) as client:
        r = await client.post(url, headers=headers, json=data)
        if r.status_code not in (200, 201):
            raise HTTPException(r.status_code, f"Supabase insert failed: {r.text[:200]}")
        rows = r.json()
        return rows[0] if rows else {}


async def _supabase_delete(
    user: UserContext, table: str, params: dict[str, str]
) -> None:
    if not supabase_configured():
        raise HTTPException(503, "Supabase not configured")
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    headers = {
        "apikey": SUPABASE_ANON_KEY,
        "Authorization": f"Bearer {user.jwt}",
    }
    async with httpx.AsyncClient(timeout=15.0) as client:
        r = await client.delete(url, headers=headers, params=params)
        if r.status_code not in (200, 204):
            raise HTTPException(r.status_code, f"Supabase delete failed: {r.text[:200]}")


# =============================================================
# Deal ↔ Broker explicit links (many-to-many)
# =============================================================

@router.post("/deal-broker-links")
async def v2_link_deal_broker(
    payload: dict,
    wc: WorkspaceContext = Depends(get_workspace_context),
):
    """Link a deal to a broker (or multiple brokers).
    payload: { deal_id, broker_id, role? }
    """
    if not wc.has("manage_deals"):
        raise HTTPException(403, "Missing permission: manage_deals")
    for k in ("deal_id", "broker_id"):
        if not payload.get(k):
            raise HTTPException(400, f"Missing {k}")
    row = await _supabase_insert_ws(
        wc,
        "deal_broker_links",
        {
            "deal_id": payload["deal_id"],
            "broker_id": payload["broker_id"],
            "role": payload.get("role", "primary"),
        },
    )
    return row


@router.get("/deal-broker-links")
async def v2_list_links(
    deal_id: str | None = None,
    broker_id: str | None = None,
    wc: WorkspaceContext = Depends(get_workspace_context),
):
    params: dict[str, str] = {"workspace_id": f"eq.{wc.workspace_id}"}
    if deal_id:
        params["deal_id"] = f"eq.{deal_id}"
    if broker_id:
        params["broker_id"] = f"eq.{broker_id}"
    rows = await _supabase_rows_ws(wc, "deal_broker_links", params=params)
    return {"links": rows}


@router.delete("/deal-broker-links/{link_id}")
async def v2_unlink(link_id: str, wc: WorkspaceContext = Depends(get_workspace_context)):
    if not wc.has("manage_deals"):
        raise HTTPException(403, "Missing permission: manage_deals")
    await _supabase_delete(
        wc.caller,
        "deal_broker_links",
        params={"id": f"eq.{link_id}", "workspace_id": f"eq.{wc.workspace_id}"},
    )
    return {"ok": True}


# =============================================================
# Meta / bootstrap
# =============================================================

@router.get("/meta")
async def v2_meta(user: UserContext = Depends(get_current_user)):
    """What's configured, what's not. Drives dashboard gating."""
    cfg = user.config
    return {
        "user_id": user.user_id,
        "email": user.email,
        "is_anonymous": user.is_anonymous,
        "wizard_completed": user.wizard_completed,
        "wizard_step": cfg.get("wizard_step", 1),
        "has_clickup": bool(cfg.get("clickup_token")),
        "has_lists": bool(user.list_deals and user.list_brokers),
        "has_gmail": bool(cfg.get("gmail_refresh_token")),
        "has_smtp": bool(cfg.get("smtp_host") and cfg.get("smtp_user")),
        "has_twilio": bool(cfg.get("twilio_account_sid")),
        "has_anthropic": bool(cfg.get("anthropic_api_key")),
        "lists": {
            "active_deals": user.list_deals,
            "brokers": user.list_brokers,
            "followups": user.list_followups,
            "touchpoints": user.list_touchpoints,
        },
    }


@router.get("/public/meta")
async def v2_public_meta():
    """Public endpoint — tells frontend if Supabase is wired.
    Used by the landing page before anyone logs in.
    """
    return {
        "supabase_configured": supabase_configured(),
        "supabase_url": SUPABASE_URL if supabase_configured() else None,
        "supabase_anon_key": SUPABASE_ANON_KEY if supabase_configured() else None,
    }


# =============================================================
# Helpers for write endpoints
# =============================================================

def _pack_helpers():
    """Lazy import pack/strip/extract so we don't bloat module load."""
    from broker_flow import pack_data, strip_data, extract_data
    return pack_data, strip_data, extract_data


async def _resolve_deals_list_id(user: UserContext) -> str:
    """Return the ClickUp list id to write new deals into.
    Priority: configured list, then first space list classified as 'deals'.
    """
    if user.list_deals:
        return user.list_deals
    lists = await _fetch_space_lists(user)
    for l in lists:
        if _classify_list(l.get("name", "")) == "deals":
            return l["id"]
    raise HTTPException(400, "No deals list found. Run the setup wizard or configure a list.")


async def _resolve_brokers_list_id(user: UserContext) -> str:
    if user.list_brokers:
        return user.list_brokers
    lists = await _fetch_space_lists(user)
    for l in lists:
        if _classify_list(l.get("name", "")) == "brokers":
            return l["id"]
    raise HTTPException(400, "No brokers list found. Run the setup wizard or configure a list.")


PRIORITY_MAP = {"urgent": 1, "high": 2, "normal": 3, "low": 4}


# =============================================================
# Deals — CRUD
# =============================================================

@router.post("/deals")
async def v2_create_deal(
    payload: dict,
    wc: WorkspaceContext = Depends(get_workspace_context),
):
    """Create a deal as a ClickUp task in the workspace owner's space.
    payload: any of {name, status, priority, tags, deal_id, asset_class, city, state,
                     units, ask_price, noi, cap_rate, broker_id, doc_status,
                     docs_received, docs_outstanding, next_action, next_action_date,
                     source, stage_entered, description_prose, list_id?}
    """
    if not wc.has("manage_deals"):
        raise HTTPException(403, "Missing permission: manage_deals")

    pack_data, _, _ = _pack_helpers()
    task_to_deal, _, _ = _lazy_mappers()
    owner = wc.owner

    list_id = payload.get("list_id") or await _resolve_deals_list_id(owner)
    name = (
        payload.get("name")
        or f"{payload.get('deal_id','JPIG-???')} · {payload.get('city','')}, {payload.get('state','')}"
    ).strip(" ·,")
    reserved = {"name", "status", "priority", "tags", "description_prose", "list_id"}
    data = {k: v for k, v in payload.items() if k not in reserved}
    prose = payload.get("description_prose", "")
    body: dict[str, Any] = {
        "name": name,
        "description": pack_data(prose, data),
    }
    if "status" in payload:
        body["status"] = payload["status"]
    if "priority" in payload:
        body["priority"] = PRIORITY_MAP.get(payload["priority"], 3)
    if "tags" in payload:
        body["tags"] = payload["tags"]

    task = await user_cu_post(owner, f"/list/{list_id}/task", body)
    deal = task_to_deal(task)
    deal["source_list_id"] = list_id
    return deal


@router.put("/deals/{task_id}")
async def v2_update_deal(
    task_id: str,
    payload: dict,
    wc: WorkspaceContext = Depends(get_workspace_context),
):
    """Patch a deal in the workspace owner's ClickUp. Merges structured fields;
    preserves prose unless overridden."""
    if not wc.has("manage_deals"):
        raise HTTPException(403, "Missing permission: manage_deals")

    pack_data, strip_data, extract_data = _pack_helpers()
    task_to_deal, _, _ = _lazy_mappers()
    owner = wc.owner

    existing = await user_cu_get(owner, f"/task/{task_id}")
    existing_desc = existing.get("description", "") or ""
    current_data = extract_data(existing_desc)

    reserved = {"name", "status", "priority", "tags", "description_prose"}
    merged = {**current_data, **{k: v for k, v in payload.items() if k not in reserved}}
    prose = payload.get("description_prose", strip_data(existing_desc))

    body: dict[str, Any] = {"description": pack_data(prose, merged)}
    if "name" in payload:
        body["name"] = payload["name"]
    if "status" in payload:
        body["status"] = payload["status"]
    if "priority" in payload:
        body["priority"] = PRIORITY_MAP.get(payload["priority"], 3)

    task = await user_cu_put(owner, f"/task/{task_id}", body)
    return task_to_deal(task)


@router.delete("/deals/{task_id}")
async def v2_delete_deal(
    task_id: str,
    wc: WorkspaceContext = Depends(get_workspace_context),
):
    if not wc.has("manage_deals"):
        raise HTTPException(403, "Missing permission: manage_deals")
    await user_cu_delete(wc.owner, f"/task/{task_id}")
    return {"ok": True}


# =============================================================
# Brokers — CRUD
# =============================================================

@router.post("/brokers")
async def v2_create_broker(
    payload: dict,
    wc: WorkspaceContext = Depends(get_workspace_context),
):
    """Create a broker record in workspace owner's ClickUp. Structured fields go in JSON block.
    payload: {name, firm?, region?, email?, phone?, relationship_strength?,
              preferred_assets?, notes?, list_id?}
    """
    if not wc.has("manage_deals"):
        raise HTTPException(403, "Missing permission: manage_deals")

    pack_data, _, _ = _pack_helpers()
    _, task_to_broker, _ = _lazy_mappers()
    owner = wc.owner

    list_id = payload.get("list_id") or await _resolve_brokers_list_id(owner)
    name = (payload.get("name") or payload.get("firm") or "Unnamed broker").strip()
    reserved = {"name", "notes", "list_id"}
    data = {k: v for k, v in payload.items() if k not in reserved}
    prose = payload.get("notes", "")
    body = {"name": name, "description": pack_data(prose, data)}

    task = await user_cu_post(owner, f"/list/{list_id}/task", body)
    broker = task_to_broker(task)
    broker["source_list_id"] = list_id
    return broker


@router.put("/brokers/{task_id}")
async def v2_update_broker(
    task_id: str,
    payload: dict,
    wc: WorkspaceContext = Depends(get_workspace_context),
):
    if not wc.has("manage_deals"):
        raise HTTPException(403, "Missing permission: manage_deals")

    pack_data, strip_data, extract_data = _pack_helpers()
    _, task_to_broker, _ = _lazy_mappers()
    owner = wc.owner

    existing = await user_cu_get(owner, f"/task/{task_id}")
    existing_desc = existing.get("description", "") or ""
    current_data = extract_data(existing_desc)

    reserved = {"name", "notes"}
    merged = {**current_data, **{k: v for k, v in payload.items() if k not in reserved}}
    prose = payload.get("notes", strip_data(existing_desc))

    body: dict[str, Any] = {"description": pack_data(prose, merged)}
    if "name" in payload:
        body["name"] = payload["name"]

    task = await user_cu_put(owner, f"/task/{task_id}", body)
    return task_to_broker(task)


@router.delete("/brokers/{task_id}")
async def v2_delete_broker(
    task_id: str,
    wc: WorkspaceContext = Depends(get_workspace_context),
):
    if not wc.has("manage_deals"):
        raise HTTPException(403, "Missing permission: manage_deals")
    await user_cu_delete(wc.owner, f"/task/{task_id}")
    return {"ok": True}


# =============================================================
# Outreach compose URLs (Gmail / Outlook / mailto)
# =============================================================

@router.post("/outreach/draft")
async def v2_outreach_draft(
    payload: dict,
    user: UserContext = Depends(get_current_user),
):
    """Build compose-deeplink URLs for Gmail, Outlook, and mailto.
    payload: {to, subject, body, cc?, bcc?}
    """
    from urllib.parse import quote
    to = (payload.get("to") or "").strip()
    subject = payload.get("subject", "")
    body = payload.get("body", "")
    cc = payload.get("cc", "")
    bcc = payload.get("bcc", "")

    q = lambda s: quote(s or "", safe="")
    mailto = f"mailto:{to}?subject={q(subject)}&body={q(body)}"
    if cc:
        mailto += f"&cc={q(cc)}"
    if bcc:
        mailto += f"&bcc={q(bcc)}"

    gmail = (
        f"https://mail.google.com/mail/?view=cm&fs=1"
        f"&to={q(to)}&su={q(subject)}&body={q(body)}"
    )
    if cc:
        gmail += f"&cc={q(cc)}"
    if bcc:
        gmail += f"&bcc={q(bcc)}"

    outlook = (
        f"https://outlook.office.com/mail/deeplink/compose"
        f"?to={q(to)}&subject={q(subject)}&body={q(body)}"
    )
    if cc:
        outlook += f"&cc={q(cc)}"

    return {"mailto": mailto, "gmail": gmail, "outlook": outlook}


# =============================================================
# Document intake — PDF parse (T-12, P&L, OM)
# =============================================================

from fastapi import File, UploadFile, Form


@router.post("/docs/parse")
async def v2_docs_parse(
    file: UploadFile = File(...),
    user: UserContext = Depends(get_current_user),
):
    """Upload a T-12 / P&L / OM PDF. Extracts text and runs heuristic regex
    to pull NOI, gross income, expenses, units, ask price, cap rate.
    Returns {filename, char_count, extracted, raw_text_preview}.
    """
    try:
        import pdfplumber
    except ImportError:
        raise HTTPException(500, "pdfplumber not installed")

    from broker_flow import heuristic_extract
    import tempfile, os as _os

    raw = await file.read()
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tf:
        tf.write(raw)
        tmp_path = tf.name

    text = ""
    try:
        with pdfplumber.open(tmp_path) as pdf:
            for page in pdf.pages:
                text += (page.extract_text() or "") + "\n"
    finally:
        try:
            _os.unlink(tmp_path)
        except Exception:
            pass

    return {
        "filename": file.filename,
        "char_count": len(text),
        "extracted": heuristic_extract(text),
        "raw_text_preview": text[:3000],
    }


# =============================================================
# Seed sample brokers (one-tap onboarding helper)
# =============================================================

@router.post("/seed/sample-brokers")
async def v2_seed_sample_brokers(wc: WorkspaceContext = Depends(get_workspace_context)):
    """Create 5 sample brokers in the workspace owner's broker list. Idempotent by name."""
    if not wc.has("manage_deals"):
        raise HTTPException(403, "Missing permission: manage_deals")

    pack_data, _, _ = _pack_helpers()
    _, task_to_broker, _ = _lazy_mappers()
    owner = wc.owner

    list_id = await _resolve_brokers_list_id(owner)

    # Check existing names to avoid dupes
    existing = await user_cu_get(
        owner, f"/list/{list_id}/task", params={"include_closed": "true"}
    )
    existing_names = {t.get("name", "").lower() for t in existing.get("tasks", []) or []}

    samples = [
        {"name": "Mark Reynolds", "firm": "Marcus & Millichap", "region": "Southwest",
         "email": "mreynolds@mmreis.com", "phone": "+1-214-555-0101",
         "relationship_strength": "Warm", "preferred_assets": ["MF", "Hotel"]},
        {"name": "Jennifer Tran", "firm": "CBRE", "region": "Southeast",
         "email": "jen.tran@cbre.com", "phone": "+1-407-555-0155",
         "relationship_strength": "Hot", "preferred_assets": ["RV", "MHP"]},
        {"name": "David Park", "firm": "JLL", "region": "Southwest",
         "email": "dpark@jll.com", "phone": "+1-915-555-0199",
         "relationship_strength": "Cold", "preferred_assets": ["Hotel", "Industrial"]},
        {"name": "Sarah Nguyen", "firm": "Newmark", "region": "Southeast",
         "email": "snguyen@newmark.com", "phone": "+1-305-555-0177",
         "relationship_strength": "Trusted", "preferred_assets": ["Self-Storage", "MF"]},
        {"name": "Tom O'Brien", "firm": "Colliers", "region": "Midwest",
         "email": "tobrien@colliers.com", "phone": "+1-312-555-0122",
         "relationship_strength": "Warm", "preferred_assets": ["Industrial", "MF"]},
    ]
    created, skipped = [], []
    for s in samples:
        if s["name"].lower() in existing_names:
            skipped.append(s["name"])
            continue
        data = {k: v for k, v in s.items() if k != "name"}
        body = {"name": s["name"], "description": pack_data("", data)}
        task = await user_cu_post(owner, f"/list/{list_id}/task", body)
        created.append(task_to_broker(task))
    return {"created": created, "skipped": skipped, "list_id": list_id}


# =============================================================
# Dashboard analytics — rollups for Reports / Heatmap / Scorecard
# =============================================================

@router.get("/analytics/summary")
async def v2_analytics_summary(wc: WorkspaceContext = Depends(get_workspace_context)):
    """Single-call payload for dashboard cards, heatmap, velocity, broker scorecard.
    Workspace-scoped — uses owner's ClickUp + caller's perms."""
    if not wc.has("view_analytics") and not wc.has("manage_deals"):
        raise HTTPException(403, "Missing permission: view_analytics")

    import time as _time
    task_to_deal, task_to_broker, _ = _lazy_mappers()

    # Pull deals + brokers via workspace-scoped list endpoints
    deals_resp = await v2_list_deals(wc)
    deals = deals_resp.get("deals", [])
    brokers_resp = await v2_list_brokers(wc)
    brokers = brokers_resp.get("brokers", [])

    # Pipeline stage counts
    stage_counts: dict[str, int] = {}
    for d in deals:
        s = (d.get("status") or "unknown").lower()
        stage_counts[s] = stage_counts.get(s, 0) + 1

    # Asset-class × stage heatmap
    heatmap: dict[str, dict[str, int]] = {}
    for d in deals:
        ac = d.get("asset_class") or "Other"
        st = (d.get("status") or "unknown").lower()
        heatmap.setdefault(ac, {})
        heatmap[ac][st] = heatmap[ac].get(st, 0) + 1

    # State-based map tally
    by_state: dict[str, int] = {}
    for d in deals:
        st = (d.get("state") or "").upper()
        if st:
            by_state[st] = by_state.get(st, 0) + 1

    # Broker scorecard: deals + complete deals + conversion
    broker_deal_count: dict[str, int] = {}
    broker_complete: dict[str, int] = {}
    for d in deals:
        bid = d.get("broker_id") or ""
        if not bid:
            continue
        broker_deal_count[bid] = broker_deal_count.get(bid, 0) + 1
        if (d.get("status") or "").lower() in ("docs complete", "underwriting", "loi", "under contract", "closed"):
            broker_complete[bid] = broker_complete.get(bid, 0) + 1

    scorecard = []
    for b in brokers:
        bid = b.get("id")
        dc = broker_deal_count.get(bid, 0)
        cc = broker_complete.get(bid, 0)
        conv = round((cc / dc) * 100, 1) if dc else 0.0
        scorecard.append({
            "broker_id": bid,
            "name": b.get("name"),
            "firm": b.get("firm"),
            "deal_count": dc,
            "complete_count": cc,
            "conversion_pct": conv,
            "score": dc * cc,
        })
    scorecard.sort(key=lambda x: x["score"], reverse=True)

    # Velocity — deals by month (date_created ms)
    velocity: dict[str, int] = {}
    for d in deals:
        dc = d.get("date_created")
        if not dc:
            continue
        try:
            ts = int(dc) / 1000
            mo = _time.strftime("%Y-%m", _time.gmtime(ts))
            velocity[mo] = velocity.get(mo, 0) + 1
        except Exception:
            pass

    return {
        "totals": {
            "deals": len(deals),
            "brokers": len(brokers),
        },
        "stage_counts": stage_counts,
        "heatmap": heatmap,
        "by_state": by_state,
        "broker_scorecard": scorecard,
        "velocity": velocity,
    }


# =============================================================
# v2 UPGRADE — deep intake parser, geocode, graph/map/reports data
# =============================================================

@router.post("/intake/deep-parse")
async def v2_intake_deep_parse(
    files: list[UploadFile] = File(default=[]),
    text: str | None = Form(default=None),
    wc: WorkspaceContext = Depends(get_workspace_context),
):
    """Advanced intake parser. Accepts PDF, XLSX, CSV, TXT, EML, DOCX or pasted
    text. Uses Claude (if workspace owner's Anthropic key set) to extract 40+
    CRE fields with confidence scores and source attribution. Falls back to
    heuristic regex extraction when no AI key is configured.
    """
    import io as _io
    import re as _re
    import json as _json

    # Use workspace owner's Anthropic key so all members share AI access
    anthropic_key = wc.owner.config.get("anthropic_api_key")

    sources: list[dict] = []
    for uf in files or []:
        name = uf.filename or "upload"
        raw = await uf.read()
        ext = (name.rsplit(".", 1)[-1].lower() if "." in name else "")

        content = ""
        kind = ext or "bin"
        err = None

        try:
            if ext == "pdf":
                import pdfplumber
                with pdfplumber.open(_io.BytesIO(raw)) as pdf:
                    content = "\n\n".join((p.extract_text() or "") for p in pdf.pages)
            elif ext in ("txt", "md", "eml", "log"):
                content = raw.decode("utf-8", errors="replace")
            elif ext == "csv":
                content = raw.decode("utf-8", errors="replace")
            elif ext in ("xlsx", "xls"):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(_io.BytesIO(raw), data_only=True)
                    parts = []
                    for sh in wb.worksheets:
                        parts.append(f"=== Sheet: {sh.title} ===")
                        for row in sh.iter_rows(values_only=True):
                            parts.append("\t".join("" if c is None else str(c) for c in row))
                    content = "\n".join(parts)
                except Exception as _e:
                    err = f"xlsx parse error: {_e}"
            elif ext in ("docx",):
                try:
                    import zipfile as _zf
                    from xml.etree import ElementTree as _ET
                    with _zf.ZipFile(_io.BytesIO(raw)) as zf:
                        xml = zf.read("word/document.xml").decode("utf-8", errors="replace")
                    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
                    root = _ET.fromstring(xml)
                    paras = ["".join(t.text or "" for t in p.findall(".//w:t", ns))
                            for p in root.findall(".//w:p", ns)]
                    content = "\n".join(paras)
                except Exception as _e:
                    err = f"docx parse error: {_e}"
            else:
                err = f"Unsupported file type: .{ext}"
        except Exception as _e:
            err = f"Extract error: {_e}"

        sources.append({
            "name": name,
            "kind": kind,
            "bytes": len(raw),
            "chars": len(content),
            "text": content,
            "error": err,
        })

    if text:
        sources.append({
            "name": "pasted-text",
            "kind": "text",
            "bytes": len(text.encode("utf-8")),
            "chars": len(text),
            "text": text,
            "error": None,
        })

    if not sources:
        raise HTTPException(400, "No files or text provided")

    total_chars = sum(s["chars"] for s in sources)
    if total_chars == 0:
        return {
            "sources": [{k: v for k, v in s.items() if k != "text"} for s in sources],
            "extracted": {},
            "mode": "empty",
            "message": "No readable text in sources.",
        }

    if not anthropic_key:
        # Heuristic-only fallback
        from broker_flow import heuristic_extract
        combined = "\n\n".join(s["text"] for s in sources if s["text"])
        basic = heuristic_extract(combined)
        # Wrap each scalar in our rich schema shape for frontend consistency
        def _wrap(v):
            return {"value": v, "confidence": 0.5 if v else 0.0, "source": "regex heuristic"}
        extracted = {
            "property": {
                "noi": _wrap(basic.get("noi")),
                "ask_price": _wrap(basic.get("ask_price")),
                "units": _wrap(basic.get("units")),
                "cap_rate": _wrap(basic.get("cap_rate")),
            },
            "financials": {
                "gross_revenue": _wrap(basic.get("gross_income")),
                "operating_expenses": _wrap(basic.get("expenses")),
            },
        }
        return {
            "sources": [{k: v for k, v in s.items() if k != "text"} for s in sources],
            "extracted": extracted,
            "mode": "heuristic",
            "message": "No Anthropic key configured. Using regex fallback. Add key in Admin → AI for deep parse.",
        }

    # Deep parse via Claude
    source_blob = "\n\n".join(
        f"=== FILE: {s['name']} ({s['kind']}, {s['chars']} chars) ==="
        f"\n{s['text'][:35000]}"
        for s in sources if s["text"]
    )
    source_blob = source_blob[:140000]  # hard cap

    schema_hint = """{
  "property": {
    "name": {"value": null, "confidence": 0, "source": ""},
    "address": {"value": null, "confidence": 0, "source": ""},
    "city": {"value": null, "confidence": 0, "source": ""},
    "state": {"value": null, "confidence": 0, "source": ""},
    "zip": {"value": null, "confidence": 0, "source": ""},
    "asset_class": {"value": null, "confidence": 0, "source": ""},
    "asset_subtype": {"value": null, "confidence": 0, "source": ""},
    "year_built": {"value": null, "confidence": 0, "source": ""},
    "year_renovated": {"value": null, "confidence": 0, "source": ""},
    "units": {"value": null, "confidence": 0, "source": ""},
    "rooms": {"value": null, "confidence": 0, "source": ""},
    "pads": {"value": null, "confidence": 0, "source": ""},
    "sqft": {"value": null, "confidence": 0, "source": ""},
    "occupancy_pct": {"value": null, "confidence": 0, "source": ""},
    "avg_rent": {"value": null, "confidence": 0, "source": ""}
  },
  "financials": {
    "ask_price": {"value": null, "confidence": 0, "source": ""},
    "noi": {"value": null, "confidence": 0, "source": ""},
    "cap_rate": {"value": null, "confidence": 0, "source": ""},
    "gross_revenue": {"value": null, "confidence": 0, "source": ""},
    "operating_expenses": {"value": null, "confidence": 0, "source": ""},
    "expense_ratio_pct": {"value": null, "confidence": 0, "source": ""},
    "price_per_unit": {"value": null, "confidence": 0, "source": ""},
    "price_per_sqft": {"value": null, "confidence": 0, "source": ""},
    "debt_in_place": {"value": null, "confidence": 0, "source": ""},
    "debt_rate_pct": {"value": null, "confidence": 0, "source": ""},
    "debt_maturity": {"value": null, "confidence": 0, "source": ""},
    "seller_financing_offered": {"value": null, "confidence": 0, "source": ""}
  },
  "market": {
    "msa": {"value": null, "confidence": 0, "source": ""},
    "submarket": {"value": null, "confidence": 0, "source": ""},
    "population_trend": {"value": null, "confidence": 0, "source": ""},
    "market_cap_rate_range": {"value": null, "confidence": 0, "source": ""}
  },
  "rent_roll": {
    "unit_mix": {"value": null, "confidence": 0, "source": ""},
    "avg_in_place_rent": {"value": null, "confidence": 0, "source": ""},
    "lease_exp_next_12mo_pct": {"value": null, "confidence": 0, "source": ""},
    "concessions": {"value": null, "confidence": 0, "source": ""}
  },
  "t12": {
    "trailing_noi_trend": {"value": null, "confidence": 0, "source": ""},
    "seasonality_notes": {"value": null, "confidence": 0, "source": ""},
    "anomalies": {"value": null, "confidence": 0, "source": ""}
  },
  "om_highlights": {
    "value_add_angle": {"value": null, "confidence": 0, "source": ""},
    "seller_motivation": {"value": null, "confidence": 0, "source": ""},
    "exclusive_listing": {"value": null, "confidence": 0, "source": ""},
    "call_for_offers_date": {"value": null, "confidence": 0, "source": ""},
    "timeline_to_close": {"value": null, "confidence": 0, "source": ""}
  },
  "broker": {
    "firm": {"value": null, "confidence": 0, "source": ""},
    "name": {"value": null, "confidence": 0, "source": ""},
    "direct_dial": {"value": null, "confidence": 0, "source": ""},
    "email": {"value": null, "confidence": 0, "source": ""},
    "license_state": {"value": null, "confidence": 0, "source": ""}
  },
  "red_flags": {
    "deferred_maintenance": {"value": null, "confidence": 0, "source": ""},
    "environmental_issues": {"value": null, "confidence": 0, "source": ""},
    "litigation": {"value": null, "confidence": 0, "source": ""},
    "unusual_clauses": {"value": null, "confidence": 0, "source": ""},
    "other_concerns": {"value": null, "confidence": 0, "source": ""}
  },
  "summary_narrative": {"value": null, "confidence": 0, "source": ""}
}"""

    prompt = (
        "You are a commercial real estate acquisitions analyst at JP Investment Group. "
        "Extract every detail from these intake documents for a new deal. "
        "For each field, return {\"value\": ..., \"confidence\": 0.0-1.0, \"source\": \"<filename>: '<exact ≤80 char quote>'\"}. "
        "Use null for fields you cannot find. Numbers should be raw numeric values (no $ or commas). "
        "Asset classes: MF, Hotel, RV, MHP, Self-Storage, Industrial, Office, Retail, Mixed-Use. "
        "summary_narrative: one-paragraph investment thesis summary.\n\n"
        "Return ONLY the JSON object, no preamble, no code fences. Schema:\n\n"
        f"{schema_hint}\n\nSOURCES:\n{source_blob}"
    )

    try:
        async with httpx.AsyncClient(timeout=90.0) as client:
            resp = await client.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": anthropic_key,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                },
                json={
                    "model": "claude-sonnet-4-5",
                    "max_tokens": 8000,
                    "messages": [{"role": "user", "content": prompt}],
                },
            )
            if resp.status_code != 200:
                raise HTTPException(502, f"Anthropic error {resp.status_code}: {resp.text[:500]}")
            data = resp.json()
            raw_out = data.get("content", [{}])[0].get("text", "")

        m = _re.search(r"\{[\s\S]*\}", raw_out)
        if not m:
            return {
                "sources": [{k: v for k, v in s.items() if k != "text"} for s in sources],
                "extracted": {},
                "mode": "anthropic",
                "error": "Model returned no JSON",
                "raw_preview": raw_out[:500],
            }
        try:
            extracted = _json.loads(m.group(0))
        except _json.JSONDecodeError as _e:
            return {
                "sources": [{k: v for k, v in s.items() if k != "text"} for s in sources],
                "extracted": {},
                "mode": "anthropic",
                "error": f"JSON decode: {_e}",
                "raw_preview": raw_out[:500],
            }
    except HTTPException:
        raise
    except Exception as _e:
        raise HTTPException(502, f"Deep parse failed: {_e}")

    return {
        "sources": [{k: v for k, v in s.items() if k != "text"} for s in sources],
        "extracted": extracted,
        "mode": "anthropic",
        "model": "claude-sonnet-4-5",
    }


async def _fetch_user_cfg(user: UserContext) -> dict:
    """Small helper — pulls user_configs row, returns dict (or empty)."""
    if not supabase_configured():
        return {}
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            r = await client.get(
                f"{SUPABASE_URL}/rest/v1/user_configs",
                params={"user_id": f"eq.{user.supabase_user_id}", "select": "*"},
                headers={
                    "apikey": SUPABASE_ANON_KEY,
                    "Authorization": f"Bearer {user.supabase_jwt}",
                },
            )
            if r.status_code == 200:
                rows = r.json()
                return rows[0] if rows else {}
    except Exception:
        pass
    return {}


@router.post("/geocode")
async def v2_geocode(payload: dict, user: UserContext = Depends(get_current_user)):
    """Geocode a single address via Nominatim (OpenStreetMap). Free, no key.
    Body: {"q": "123 Main St, Austin, TX"}. Returns {lat, lng, display_name}.
    """
    q = (payload or {}).get("q") or ""
    q = q.strip()
    if not q:
        raise HTTPException(400, "Address required")
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            r = await client.get(
                "https://nominatim.openstreetmap.org/search",
                params={"q": q, "format": "json", "limit": 1, "addressdetails": 1},
                headers={"User-Agent": "JP-BrokerFlow/2.0 (acquisitions tool)"},
            )
            if r.status_code != 200:
                return {"error": f"Nominatim {r.status_code}", "q": q}
            arr = r.json()
            if not arr:
                return {"lat": None, "lng": None, "q": q, "error": "Not found"}
            hit = arr[0]
            return {
                "q": q,
                "lat": float(hit.get("lat", 0)),
                "lng": float(hit.get("lon", 0)),
                "display_name": hit.get("display_name"),
                "country": (hit.get("address") or {}).get("country"),
                "state": (hit.get("address") or {}).get("state"),
                "city": (hit.get("address") or {}).get("city") or (hit.get("address") or {}).get("town"),
            }
    except Exception as _e:
        return {"error": str(_e), "q": q}


@router.get("/analytics/graph")
async def v2_analytics_graph(wc: WorkspaceContext = Depends(get_workspace_context)):
    """Enriched network graph — brokers, deals, markets, asset classes.
    Workspace-scoped. Edges: broker↔deal, deal↔market, broker↔asset-class,
    broker↔broker (co-market). Each node carries rich metadata for inspector.
    """
    if not wc.has("view_analytics") and not wc.has("manage_deals"):
        raise HTTPException(403, "Missing permission: view_analytics")

    import datetime as _dt
    import time as _time

    deals_resp = await v2_list_deals(wc)
    brokers_resp = await v2_list_brokers(wc)
    deals = deals_resp.get("deals", [])
    brokers = brokers_resp.get("brokers", [])

    # Touchpoints — count + last-touched per (deal,broker) and per broker
    tp_counts: dict[tuple, int] = {}
    last_touch_per_broker: dict[str, str] = {}
    try:
        tp_resp = await v2_query_touchpoints(wc=wc)
        for tp in tp_resp.get("touchpoints", []):
            did = tp.get("deal_id") or ""
            bid = tp.get("broker_id") or ""
            if did and bid:
                key = (did, bid)
                tp_counts[key] = tp_counts.get(key, 0) + 1
            ts = tp.get("sent_at") or tp.get("created_at")
            if bid and ts:
                prev = last_touch_per_broker.get(bid)
                if not prev or ts > prev:
                    last_touch_per_broker[bid] = ts
    except Exception:
        pass

    now_ms = int(_time.time() * 1000)
    _ICP_ASSETS = {"Multifamily", "MF", "Hotel", "RV", "MHP", "Self-Storage", "Industrial"}

    def _days_since(iso_ts: str | None) -> int | None:
        if not iso_ts:
            return None
        try:
            dt = _dt.datetime.fromisoformat(iso_ts.replace("Z", "+00:00"))
            return max(0, (_dt.datetime.now(_dt.timezone.utc) - dt).days)
        except Exception:
            return None

    # Index deals by broker + build broker→asset-class counts
    deals_by_broker: dict[str, list] = {}
    broker_asset_mix: dict[str, dict[str, int]] = {}
    broker_states: dict[str, set] = {}
    for d in deals:
        bid = d.get("broker_id") or ""
        if not bid:
            continue
        deals_by_broker.setdefault(bid, []).append(d)
        ac = d.get("asset_class") or "Other"
        broker_asset_mix.setdefault(bid, {})
        broker_asset_mix[bid][ac] = broker_asset_mix[bid].get(ac, 0) + 1
        st = (d.get("state") or "").upper()
        if st:
            broker_states.setdefault(bid, set()).add(st)

    nodes = []

    # Broker nodes — enriched
    for b in brokers:
        bid = b["id"]
        bdeals = deals_by_broker.get(bid, [])
        closed = sum(1 for d in bdeals if (d.get("status") or "").lower() == "closed")
        active = sum(1 for d in bdeals if (d.get("status") or "").lower() not in ("closed", "dead"))
        last_ts = last_touch_per_broker.get(bid)
        asks = [float(d.get("ask_price") or 0) for d in bdeals if d.get("ask_price")]
        avg_ask = (sum(asks) / len(asks)) if asks else 0
        icp_hits = sum(1 for d in bdeals if (d.get("asset_class") or "") in _ICP_ASSETS)
        nodes.append({
            "id": f"b:{bid}",
            "type": "broker",
            "label": b.get("name") or "",
            "sublabel": b.get("firm") or "",
            "group": b.get("relationship_strength") or "Cold",
            "value": max(1, len(bdeals)),
            "meta": {
                "firm": b.get("firm"),
                "tier": b.get("relationship_strength") or "Cold",
                "total_deals": len(bdeals),
                "closed": closed,
                "active": active,
                "avg_ask": round(avg_ask, 0),
                "last_touch": last_ts,
                "days_since_touch": _days_since(last_ts),
                "icp_hits": icp_hits,
                "asset_mix": broker_asset_mix.get(bid, {}),
            },
        })

    # Deal nodes — enriched
    markets: dict[str, int] = {}
    market_value: dict[str, float] = {}
    for d in deals:
        did = d["id"]
        ac = d.get("asset_class") or "Other"
        age = None
        if d.get("date_created"):
            try:
                age = int((now_ms - int(d["date_created"])) / (1000 * 86400))
            except Exception:
                age = None
        nodes.append({
            "id": f"d:{did}",
            "type": "deal",
            "label": d.get("name") or "",
            "sublabel": d.get("status") or "",
            "group": (d.get("status") or "unknown").lower(),
            "value": 1,
            "meta": {
                "status": d.get("status"),
                "asset_class": ac,
                "state": d.get("state"),
                "city": d.get("city"),
                "ask_price": d.get("ask_price"),
                "noi": d.get("noi"),
                "cap_rate": d.get("cap_rate"),
                "units": d.get("units"),
                "age_days": age,
                "broker_id": d.get("broker_id"),
            },
        })
        mk = (d.get("state") or "").upper()
        if mk:
            markets[mk] = markets.get(mk, 0) + 1
            try:
                market_value[mk] = market_value.get(mk, 0) + float(d.get("ask_price") or 0)
            except Exception:
                pass

    # Market nodes — enriched
    for mk, n in markets.items():
        nodes.append({
            "id": f"m:{mk}",
            "type": "market",
            "label": mk,
            "sublabel": f"{n} deal(s)",
            "group": "market",
            "value": max(1, n),
            "meta": {"deal_count": n, "total_value": round(market_value.get(mk, 0), 0)},
        })

    # Asset-class nodes — one per observed class (hubs)
    asset_counts: dict[str, int] = {}
    for d in deals:
        ac = d.get("asset_class") or "Other"
        asset_counts[ac] = asset_counts.get(ac, 0) + 1
    for ac, n in asset_counts.items():
        nodes.append({
            "id": f"a:{ac}",
            "type": "asset",
            "label": ac,
            "sublabel": f"{n} deal(s)",
            "group": "asset",
            "value": max(1, n),
            "meta": {"deal_count": n, "icp": ac in _ICP_ASSETS},
        })

    links = []
    # broker ↔ deal
    for d in deals:
        bid = d.get("broker_id")
        if bid:
            links.append({
                "source": f"b:{bid}",
                "target": f"d:{d['id']}",
                "weight": 1 + tp_counts.get((d["id"], bid), 0),
                "kind": "broker-deal",
            })
        mk = (d.get("state") or "").upper()
        if mk:
            links.append({
                "source": f"d:{d['id']}",
                "target": f"m:{mk}",
                "weight": 1,
                "kind": "deal-market",
            })

    # broker ↔ asset-class (aggregated from historical deals)
    for bid, mix in broker_asset_mix.items():
        for ac, cnt in mix.items():
            links.append({
                "source": f"b:{bid}",
                "target": f"a:{ac}",
                "weight": cnt,
                "kind": "broker-asset",
            })

    # broker ↔ broker co-market — two brokers sharing same state(s)
    b_list = list(broker_states.items())
    for i in range(len(b_list)):
        for j in range(i + 1, len(b_list)):
            overlap = b_list[i][1] & b_list[j][1]
            if overlap:
                links.append({
                    "source": f"b:{b_list[i][0]}",
                    "target": f"b:{b_list[j][0]}",
                    "weight": len(overlap),
                    "kind": "broker-broker",
                    "shared_markets": sorted(list(overlap)),
                })

    return {
        "nodes": nodes,
        "links": links,
        "counts": {
            "brokers": len(brokers),
            "deals": len(deals),
            "markets": len(markets),
            "assets": len(asset_counts),
            "edges": len(links),
        },
    }


@router.get("/analytics/heatmap-activity")
async def v2_analytics_heatmap_activity(wc: WorkspaceContext = Depends(get_workspace_context)):
    """Broker × week activity heatmap — touchpoints per broker per ISO week
    (workspace-scoped)."""
    if not wc.has("view_analytics") and not wc.has("manage_deals"):
        raise HTTPException(403, "Missing permission: view_analytics")

    import datetime as _dt

    tp_resp = await v2_query_touchpoints(wc=wc)
    tps = tp_resp.get("touchpoints", [])
    brokers_resp = await v2_list_brokers(wc)
    brokers = brokers_resp.get("brokers", [])
    bmap = {b["id"]: b for b in brokers}

    weeks = []
    today = _dt.date.today()
    monday = today - _dt.timedelta(days=today.weekday())
    for i in range(11, -1, -1):  # last 12 weeks
        wk = monday - _dt.timedelta(weeks=i)
        weeks.append(wk.isoformat())

    grid: dict[str, dict[str, int]] = {b["id"]: {w: 0 for w in weeks} for b in brokers}
    for tp in tps:
        bid = tp.get("broker_id")
        ts = tp.get("sent_at") or tp.get("created_at")
        if not bid or not ts or bid not in grid:
            continue
        try:
            dt = _dt.datetime.fromisoformat(ts.replace("Z", "+00:00"))
            wk_monday = (dt.date() - _dt.timedelta(days=dt.date().weekday())).isoformat()
            if wk_monday in grid[bid]:
                grid[bid][wk_monday] = grid[bid][wk_monday] + 1
        except Exception:
            continue

    rows = []
    for b in brokers:
        bid = b["id"]
        row = {
            "broker_id": bid,
            "name": b.get("name"),
            "firm": b.get("firm"),
            "tier": b.get("relationship_strength") or "Cold",
            "cells": [grid[bid][w] for w in weeks],
            "total": sum(grid[bid].values()),
        }
        rows.append(row)
    rows.sort(key=lambda r: r["total"], reverse=True)
    return {"weeks": weeks, "rows": rows}


@router.get("/analytics/reports")
async def v2_analytics_reports(wc: WorkspaceContext = Depends(get_workspace_context)):
    """Extended reporting rollup: stage funnel, avg days per stage, broker
    leaderboard, source attribution, cadence conversion rate (workspace-scoped).
    """
    if not wc.has("view_analytics") and not wc.has("manage_deals"):
        raise HTTPException(403, "Missing permission: view_analytics")

    import time as _time
    import datetime as _dt

    deals_resp = await v2_list_deals(wc)
    deals = deals_resp.get("deals", [])
    brokers_resp = await v2_list_brokers(wc)
    brokers = brokers_resp.get("brokers", [])
    bmap = {b["id"]: b for b in brokers}

    stages = ["incoming leads", "docs requested", "docs complete", "underwriting", "loi", "under contract", "closed", "dead"]
    funnel_counts = {s: 0 for s in stages}
    for d in deals:
        s = (d.get("status") or "").lower()
        if s in funnel_counts:
            funnel_counts[s] = funnel_counts[s] + 1

    # Build ordered array w/ stage-over-stage conversion %
    funnel = []
    _prev = None
    for _s in stages:
        _n = funnel_counts[_s]
        _conv = 100.0 if _prev is None else (round((_n / _prev) * 100, 1) if _prev else 0.0)
        funnel.append({"stage": _s, "count": _n, "conversion_pct": _conv})
        _prev = _n

    # Avg time in stage — rough, from date_created to now for non-terminal deals
    now_ms = int(_time.time() * 1000)
    days_in_stage: dict[str, list[int]] = {s: [] for s in stages}
    for d in deals:
        dc = d.get("date_created")
        s = (d.get("status") or "").lower()
        if not dc or s not in days_in_stage:
            continue
        try:
            age = (now_ms - int(dc)) / (1000 * 86400)
            days_in_stage[s].append(age)
        except Exception:
            pass
    avg_days = {s: (round(sum(v) / len(v), 1) if v else 0) for s, v in days_in_stage.items()}

    # Broker leaderboard — richer + multi-factor relationship score (0-100)
    import math as _math

    # Pull touchpoints once for recency + reply-rate calc
    tp_by_broker: dict[str, list] = {}
    try:
        _tp_all = await v2_query_touchpoints(wc=wc)
        for _tp in _tp_all.get("touchpoints", []):
            _bid = _tp.get("broker_id")
            if _bid:
                tp_by_broker.setdefault(_bid, []).append(_tp)
    except Exception:
        pass

    _ICP_ASSETS = {"Multifamily", "MF", "Hotel", "RV", "MHP", "Self-Storage", "Industrial"}
    _TIER_BONUS = {"Hot": 5, "Trusted": 4, "Warm": 3, "New": 2, "Cold": 1}

    def _days_since_iso(ts: str | None) -> float:
        if not ts:
            return 999.0
        try:
            dt = _dt.datetime.fromisoformat(ts.replace("Z", "+00:00"))
            return max(0.0, (_dt.datetime.now(_dt.timezone.utc) - dt).total_seconds() / 86400.0)
        except Exception:
            return 999.0

    leaderboard = []
    for b in brokers:
        bid = b["id"]
        bdeals = [d for d in deals if d.get("broker_id") == bid]
        n_deals = len(bdeals)
        closed = sum(1 for d in bdeals if (d.get("status") or "").lower() == "closed")
        active = sum(1 for d in bdeals if (d.get("status") or "").lower() not in ("closed", "dead"))
        dead = sum(1 for d in bdeals if (d.get("status") or "").lower() == "dead")
        docs_in = sum(1 for d in bdeals if (d.get("status") or "").lower() in ("docs complete", "underwriting", "loi", "under contract", "closed"))

        # Sub-score 1: volume (0-25) — log-scaled deal count
        volume_score = round(min(25.0, 10.0 * _math.log1p(n_deals)), 1)

        # Sub-score 2: quality (0-25) — close_rate × closed-count bonus
        close_rate = (closed / n_deals) if n_deals else 0.0
        quality_score = round(min(25.0, close_rate * 50.0 + closed * 5.0), 1)

        # Sub-score 3: engagement (0-20) — docs-in rate + touchpoint volume (log)
        docs_in_rate = (docs_in / n_deals) if n_deals else 0.0
        tp_count = len(tp_by_broker.get(bid, []))
        engagement_score = round(min(20.0, docs_in_rate * 12.0 + min(8.0, 2.0 * _math.log1p(tp_count))), 1)

        # Sub-score 4: ICP alignment (0-15) — JPIG-preferred asset classes
        icp_hits = sum(1 for d in bdeals if (d.get("asset_class") or "") in _ICP_ASSETS)
        alignment_ratio = (icp_hits / n_deals) if n_deals else 0.0
        alignment_score = round(min(15.0, alignment_ratio * 15.0), 1)

        # Sub-score 5: recency (0-10) — exp decay on days since last touchpoint
        last_ts = None
        for _tp in tp_by_broker.get(bid, []):
            _ts = _tp.get("sent_at") or _tp.get("created_at")
            if _ts and (not last_ts or _ts > last_ts):
                last_ts = _ts
        days_since = _days_since_iso(last_ts)
        # Full score inside 7 days, decay to ~0 by 90 days
        recency_score = round(10.0 * _math.exp(-days_since / 30.0), 1) if days_since < 999 else 0.0

        # Sub-score 6: tier bonus (0-5)
        tier = b.get("relationship_strength") or "Cold"
        tier_bonus = float(_TIER_BONUS.get(tier, 1))

        total_score = round(volume_score + quality_score + engagement_score + alignment_score + recency_score + tier_bonus, 1)

        # Avg deal size for reference
        asks = [float(d.get("ask_price") or 0) for d in bdeals if d.get("ask_price")]
        avg_ask = round((sum(asks) / len(asks)), 0) if asks else 0

        leaderboard.append({
            "broker_id": bid,
            "name": b.get("name"),
            "firm": b.get("firm"),
            "tier": tier,
            "total": n_deals,
            "active": active,
            "closed": closed,
            "dead": dead,
            "total_deals": n_deals,
            "active_deals": active,
            "closed_deals": closed,
            "docs_in_rate": round(docs_in_rate * 100, 1),
            "close_rate": round(close_rate * 100, 1),
            "icp_hits": icp_hits,
            "icp_alignment_pct": round(alignment_ratio * 100, 1),
            "avg_ask": avg_ask,
            "touchpoints": tp_count,
            "days_since_touch": round(days_since, 1) if days_since < 999 else None,
            "score": total_score,
            "score_breakdown": {
                "volume": volume_score,
                "quality": quality_score,
                "engagement": engagement_score,
                "alignment": alignment_score,
                "recency": recency_score,
                "tier_bonus": tier_bonus,
            },
        })
    leaderboard.sort(key=lambda x: x["score"], reverse=True)

    # Source attribution — asset class mix per broker tier
    tier_x_asset: dict[str, dict[str, int]] = {}
    for d in deals:
        bid = d.get("broker_id")
        b = bmap.get(bid) if bid else None
        tier = (b.get("relationship_strength") if b else None) or "Unknown"
        ac = d.get("asset_class") or "Other"
        tier_x_asset.setdefault(tier, {})
        tier_x_asset[tier][ac] = tier_x_asset[tier].get(ac, 0) + 1

    # Cadence conversion — deals moved past "incoming leads" vs total
    total = len(deals)
    moved = sum(1 for d in deals if (d.get("status") or "").lower() not in ("incoming leads", "dead"))
    cadence_conv = round((moved / total) * 100, 1) if total else 0

    return {
        "funnel": funnel,
        "avg_days_in_stage": avg_days,
        "leaderboard": leaderboard,
        "tier_x_asset": tier_x_asset,
        "cadence_conversion_pct": cadence_conv,
        "totals": {"deals": total, "brokers": len(brokers), "moved_past_incoming": moved},
    }


# =============================================================================
# USER CONFIG — save integration keys (writes to Supabase user_configs table)
# =============================================================================

def _mask(val: str | None) -> str | None:
    if not val:
        return None
    s = str(val)
    if len(s) <= 8:
        return "••••"
    return s[:4] + "••••" + s[-4:]


@router.get("/user-config")
async def v2_get_user_config(user: UserContext = Depends(get_current_user)):
    """Returns current user config with sensitive fields masked."""
    cfg = await _fetch_user_cfg(user)
    masked = dict(cfg)
    for k in ("clickup_token", "smtp_password", "twilio_token", "twilio_sid", "anthropic_api_key"):
        if masked.get(k):
            masked[f"{k}_mask"] = _mask(masked.pop(k))
    return masked


@router.put("/user-config")
async def v2_put_user_config(payload: dict, user: UserContext = Depends(get_current_user)):
    """Merge-update user_configs row for the current user.

    Only fields present in the payload are touched. Secrets containing '•' are
    treated as 'unchanged' (the client is sending back the mask) and skipped.
    """
    if not supabase_configured():
        raise HTTPException(500, "Supabase not configured")

    clean: dict = {}
    for k, v in (payload or {}).items():
        if v is None:
            continue
        if isinstance(v, str) and "•" in v:
            continue  # skip masked values
        clean[k] = v

    if not clean:
        return {"ok": True, "updated": 0}

    clean["user_id"] = user.supabase_user_id

    # Upsert via PostgREST. Conflict on user_id; prefer-resolution=merge-duplicates.
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            r = await client.post(
                f"{SUPABASE_URL}/rest/v1/user_configs",
                params={"on_conflict": "user_id"},
                headers={
                    "apikey": SUPABASE_ANON_KEY,
                    "Authorization": f"Bearer {user.supabase_jwt}",
                    "Content-Type": "application/json",
                    "Prefer": "resolution=merge-duplicates,return=representation",
                },
                json=clean,
            )
            if r.status_code not in (200, 201):
                raise HTTPException(r.status_code, f"Supabase error: {r.text[:300]}")
            return {"ok": True, "updated": len(clean), "row": (r.json() or [None])[0]}
    except HTTPException:
        raise
    except Exception as _e:
        raise HTTPException(502, f"user-config write failed: {_e}")


# =============================================================================
# ADMIN — health, integration tests, danger zone, team invites
# =============================================================================

@router.get("/admin/health")
async def v2_admin_health(user: UserContext = Depends(get_current_user)):
    """Lightweight env health — reports which integrations are wired up per user."""
    cfg = await _fetch_user_cfg(user)
    status: dict[str, str] = {}
    status["supabase"] = "ok" if supabase_configured() else "fail"
    status["clickup"] = "ok" if cfg.get("clickup_token") else "unknown"
    status["smtp"] = "ok" if cfg.get("smtp_password") and cfg.get("smtp_host") else "unknown"
    status["twilio"] = "ok" if cfg.get("twilio_sid") and cfg.get("twilio_token") else "unknown"
    status["anthropic"] = "ok" if cfg.get("anthropic_api_key") else "unknown"
    status["nominatim"] = "ok"  # always available
    return status


@router.post("/admin/test")
async def v2_admin_test(payload: dict, user: UserContext = Depends(get_current_user)):
    """Run a live integration test for the named kind.
    Body: {"kind": "clickup" | "smtp" | "sms" | "anthropic"}
    """
    kind = (payload or {}).get("kind")
    cfg = await _fetch_user_cfg(user)
    try:
        if kind == "clickup":
            tok = cfg.get("clickup_token")
            if not tok:
                return {"ok": False, "error": "no token saved"}
            async with httpx.AsyncClient(timeout=15.0) as client:
                r = await client.get("https://api.clickup.com/api/v2/user", headers={"Authorization": tok})
                if r.status_code == 200:
                    return {"ok": True, "detail": r.json().get("user", {}).get("username", "ok")}
                return {"ok": False, "error": f"clickup {r.status_code}"}
        if kind == "anthropic":
            key = cfg.get("anthropic_api_key")
            if not key:
                return {"ok": False, "error": "no key saved"}
            async with httpx.AsyncClient(timeout=20.0) as client:
                r = await client.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={"x-api-key": key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                    json={"model": cfg.get("anthropic_model") or "claude-sonnet-4-5", "max_tokens": 10, "messages": [{"role": "user", "content": "ping"}]},
                )
                return {"ok": r.status_code == 200, "error": None if r.status_code == 200 else f"anthropic {r.status_code}"}
        if kind == "smtp":
            if not cfg.get("smtp_host"):
                return {"ok": False, "error": "no smtp host"}
            # Light TCP-level check — don't send an actual email here.
            import socket
            try:
                with socket.create_connection((cfg["smtp_host"], int(cfg.get("smtp_port") or 587)), timeout=6) as s:
                    return {"ok": True, "detail": "smtp reachable"}
            except Exception as _e:
                return {"ok": False, "error": str(_e)[:120]}
        if kind == "sms":
            if not cfg.get("twilio_sid") or not cfg.get("twilio_token"):
                return {"ok": False, "error": "twilio creds missing"}
            async with httpx.AsyncClient(timeout=15.0, auth=(cfg["twilio_sid"], cfg["twilio_token"])) as client:
                r = await client.get(f"https://api.twilio.com/2010-04-01/Accounts/{cfg['twilio_sid']}.json")
                return {"ok": r.status_code == 200, "error": None if r.status_code == 200 else f"twilio {r.status_code}"}
        return {"ok": False, "error": f"unknown kind: {kind}"}
    except Exception as _e:
        return {"ok": False, "error": str(_e)[:200]}


@router.post("/admin/danger")
async def v2_admin_danger(payload: dict, user: UserContext = Depends(get_current_user)):
    """Danger-zone actions. Body: {"action": "revoke-clickup" | "clear-analytics"}"""
    action = (payload or {}).get("action")
    if action == "revoke-clickup":
        # Overwrite the ClickUp token with NULL.
        try:
            async with httpx.AsyncClient(timeout=15.0) as client:
                r = await client.patch(
                    f"{SUPABASE_URL}/rest/v1/user_configs",
                    params={"user_id": f"eq.{user.supabase_user_id}"},
                    headers={
                        "apikey": SUPABASE_ANON_KEY,
                        "Authorization": f"Bearer {user.supabase_jwt}",
                        "Content-Type": "application/json",
                        "Prefer": "return=minimal",
                    },
                    json={"clickup_token": None, "clickup_space_id": None},
                )
                return {"ok": r.status_code in (200, 204), "detail": "clickup revoked"}
        except Exception as _e:
            return {"ok": False, "error": str(_e)}
    if action == "clear-analytics":
        # No server-side cache to clear in this deploy — signal the client.
        return {"ok": True, "detail": "client should refresh"}
    return {"ok": False, "error": f"unknown action: {action}"}


@router.post("/admin/invite")
async def v2_admin_invite(payload: dict, user: UserContext = Depends(get_current_user)):
    """Legacy single-team invite stub. See /teams/{team_id}/invitations for the
    real multi-tenant flow below. Kept for backward compatibility with the old
    Admin view.
    """
    email = (payload or {}).get("email")
    role = (payload or {}).get("role") or "Analyst"
    if not email or "@" not in email:
        raise HTTPException(400, "valid email required")
    return {"ok": True, "email": email, "role": role, "status": "queued"}


# =============================================================
# TEAMS / WORKSPACES
# Multi-tenant collaboration: owner creates team, invites members by
# email, assigns role (admin, member, viewer, or custom). Backend
# enforces every permission check — RLS on these tables is read-only
# (members can see their team). Writes use the service-role key so
# the backend is the single source of truth for auth decisions.
# =============================================================

SUPABASE_SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "").strip()


def _require_service_role() -> None:
    if not SUPABASE_SERVICE_ROLE_KEY:
        raise HTTPException(
            503,
            "SUPABASE_SERVICE_ROLE_KEY env var not set — teams API unavailable. "
            "Add it in Render → Environment and redeploy.",
        )


async def _svc_request(
    method: str,
    path: str,
    params: dict[str, str] | None = None,
    body: Any = None,
    prefer: str = "return=representation",
) -> Any:
    """HTTP call against Supabase REST using the service-role key.
    Bypasses RLS — only called from backend after permission checks.
    """
    _require_service_role()
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "Content-Type": "application/json",
        "Accept": "application/json",
        "Prefer": prefer,
    }
    async with httpx.AsyncClient(timeout=15.0) as client:
        r = await client.request(
            method, url, headers=headers, params=params or {}, json=body
        )
        if r.status_code not in (200, 201, 204):
            raise HTTPException(
                r.status_code, f"Supabase svc call failed: {r.text[:300]}"
            )
        if r.status_code == 204 or not r.text:
            return None
        try:
            return r.json()
        except Exception:
            return None


async def _svc_select(path: str, params: dict[str, str]) -> list[dict]:
    rows = await _svc_request("GET", path, params=params)
    return rows if isinstance(rows, list) else []


async def _svc_insert(path: str, body: dict | list[dict]) -> list[dict]:
    rows = await _svc_request("POST", path, body=body)
    return rows if isinstance(rows, list) else ([rows] if rows else [])


async def _svc_update(
    path: str, params: dict[str, str], body: dict
) -> list[dict]:
    rows = await _svc_request("PATCH", path, params=params, body=body)
    return rows if isinstance(rows, list) else ([rows] if rows else [])


async def _svc_delete(path: str, params: dict[str, str]) -> None:
    await _svc_request("DELETE", path, params=params, prefer="return=minimal")


# ---------- Supabase auth admin helpers (identity enrichment) ----------

async def _svc_auth_admin_get_user(user_id: str) -> dict | None:
    """Fetch a single auth.users row via the Supabase admin API.
    Needs service-role key. Returns None on 404.
    """
    _require_service_role()
    url = f"{SUPABASE_URL}/auth/v1/admin/users/{user_id}"
    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "Accept": "application/json",
    }
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            r = await client.get(url, headers=headers)
            if r.status_code == 404:
                return None
            if r.status_code != 200:
                return None
            return r.json()
    except Exception:
        return None


def _identity_from_auth_user(u: dict | None) -> dict:
    if not u:
        return {"email": None, "display_name": None, "avatar_url": None}
    meta = u.get("user_metadata") or {}
    email = u.get("email")
    display = (
        meta.get("display_name")
        or meta.get("full_name")
        or meta.get("name")
        or (email.split("@")[0] if email else None)
    )
    return {
        "email": email,
        "display_name": display,
        "avatar_url": meta.get("avatar_url"),
    }


async def _identity_map(user_ids: list[str]) -> dict[str, dict]:
    """Return {user_id: {email, display_name, avatar_url}} for the given ids.
    Deduplicates, runs admin lookups in parallel, swallows failures per user.
    """
    import asyncio

    uniq = [u for u in dict.fromkeys(user_ids) if u]
    if not uniq or not SUPABASE_SERVICE_ROLE_KEY:
        return {u: _identity_from_auth_user(None) for u in uniq}
    results = await asyncio.gather(
        *[_svc_auth_admin_get_user(u) for u in uniq], return_exceptions=True
    )
    out: dict[str, dict] = {}
    for uid, res in zip(uniq, results):
        if isinstance(res, Exception):
            out[uid] = _identity_from_auth_user(None)
        else:
            out[uid] = _identity_from_auth_user(res)
    return out


async def _enrich_members_with_identity(rows: list[dict]) -> list[dict]:
    if not rows:
        return rows
    ids = [r.get("user_id") for r in rows if r.get("user_id")]
    idmap = await _identity_map(ids)
    for r in rows:
        ident = idmap.get(r.get("user_id")) or _identity_from_auth_user(None)
        r["email"] = ident["email"]
        r["display_name"] = ident["display_name"]
        r["avatar_url"] = ident["avatar_url"]
    return rows


# ---------- Permission helpers ----------

# =================================================================
# Unified role catalog — one list per workspace held in public.team_roles.
# Each role carries BOTH scopes of permissions:
#   * permissions      — workspace-scope (admin/manage_members/manage_roles/
#                       manage_deals/view_analytics)
#   * team_permissions — sub-team-scope (team_admin/manage_team_members/
#                       manage_team_roles/edit_team_content/view_team_content)
# Mirror rows in public.workspace_team_roles are kept in sync by
# _mirror_catalog_role_to_subteams / _unmirror_catalog_role on every write
# so sub-team dropdowns always show the same role list as the workspace.
# =================================================================

DEFAULT_ROLES = [
    {
        "name": "Owner",
        "is_system": True,
        "is_default": False,
        "permissions": {
            "admin": True,
            "manage_members": True,
            "manage_roles": True,
            "manage_deals": True,
            "view_analytics": True,
        },
        "team_permissions": {
            "team_admin": True,
            "manage_team_members": True,
            "manage_team_roles": True,
            "edit_team_content": True,
            "view_team_content": True,
        },
    },
    {
        "name": "Admin",
        "is_system": False,
        "is_default": True,
        "permissions": {
            "admin": True,
            "manage_members": True,
            "manage_roles": True,
            "manage_deals": True,
            "view_analytics": True,
        },
        "team_permissions": {
            "team_admin": True,
            "manage_team_members": True,
            "manage_team_roles": True,
            "edit_team_content": True,
            "view_team_content": True,
        },
    },
    {
        "name": "Team Admin",
        "is_system": True,
        "is_default": False,
        "permissions": {
            "admin": False,
            "manage_members": False,
            "manage_roles": False,
            "manage_deals": False,
            "view_analytics": False,
        },
        "team_permissions": {
            "team_admin": True,
            "manage_team_members": True,
            "manage_team_roles": True,
            "edit_team_content": True,
            "view_team_content": True,
        },
    },
    {
        "name": "Role 1",
        "is_system": False,
        "is_default": False,
        "permissions": {
            "admin": False,
            "manage_members": False,
            "manage_roles": False,
            "manage_deals": False,
            "view_analytics": False,
        },
        "team_permissions": {
            "team_admin": False,
            "manage_team_members": False,
            "manage_team_roles": False,
            "edit_team_content": False,
            "view_team_content": False,
        },
    },
    {
        "name": "Role 2",
        "is_system": False,
        "is_default": False,
        "permissions": {
            "admin": False,
            "manage_members": False,
            "manage_roles": False,
            "manage_deals": False,
            "view_analytics": False,
        },
        "team_permissions": {
            "team_admin": False,
            "manage_team_members": False,
            "manage_team_roles": False,
            "edit_team_content": False,
            "view_team_content": False,
        },
    },
    {
        "name": "Role 3",
        "is_system": False,
        "is_default": False,
        "permissions": {
            "admin": False,
            "manage_members": False,
            "manage_roles": False,
            "manage_deals": False,
            "view_analytics": False,
        },
        "team_permissions": {
            "team_admin": False,
            "manage_team_members": False,
            "manage_team_roles": False,
            "edit_team_content": False,
            "view_team_content": False,
        },
    },
]


async def _get_team_or_404(team_id: str) -> dict:
    rows = await _svc_select(
        "teams", {"id": f"eq.{team_id}", "select": "*", "limit": "1"}
    )
    if not rows:
        raise HTTPException(404, "Team not found")
    return rows[0]


async def _get_membership(team_id: str, user_id: str) -> dict | None:
    rows = await _svc_select(
        "team_members",
        {
            "team_id": f"eq.{team_id}",
            "user_id": f"eq.{user_id}",
            "select": "*,team_roles(*)",
            "limit": "1",
        },
    )
    return rows[0] if rows else None


async def _require_member(team_id: str, user: UserContext) -> tuple[dict, dict]:
    """Returns (team, membership). Raises 403 if user is not a member."""
    team = await _get_team_or_404(team_id)
    if team["owner_id"] == user.user_id:
        # Owner is always a member; synthesize membership if row missing.
        m = await _get_membership(team_id, user.user_id)
        return team, (m or {"role_id": None, "team_roles": None, "is_owner": True})
    m = await _get_membership(team_id, user.user_id)
    if not m:
        raise HTTPException(403, "Not a member of this team")
    return team, m


def _has_perm(team: dict, membership: dict, user: UserContext, perm: str) -> bool:
    if team["owner_id"] == user.user_id:
        return True
    role = (membership or {}).get("team_roles") or {}
    perms = role.get("permissions") or {}
    return bool(perms.get("admin") or perms.get(perm))


async def _require_perm(
    team_id: str, user: UserContext, perm: str
) -> tuple[dict, dict]:
    team, m = await _require_member(team_id, user)
    if not _has_perm(team, m, user, perm):
        raise HTTPException(403, f"Missing permission: {perm}")
    return team, m


async def _ensure_default_roles(team_id: str) -> list[dict]:
    """Ensure every workspace has the six system/default catalog roles
    (Owner, Admin, Team Admin, Role 1/2/3). Idempotent — inserts any missing
    rows; existing rows are left intact (including user customizations).
    Each catalog row carries BOTH workspace-scope permissions and
    sub-team-scope team_permissions. After ensuring the catalog, mirror
    every role into every existing sub-team so sub-team dropdowns match.
    """
    existing = await _svc_select(
        "team_roles", {"team_id": f"eq.{team_id}", "select": "*"}
    )
    existing_names = {r["name"] for r in existing}
    to_insert = [
        {
            "team_id": team_id,
            "name": r["name"],
            "permissions": r["permissions"],
            "team_permissions": r["team_permissions"],
            "is_default": r["is_default"],
            "is_system": r["is_system"],
        }
        for r in DEFAULT_ROLES
        if r["name"] not in existing_names
    ]
    if to_insert:
        inserted = await _svc_insert("team_roles", to_insert)
        existing = existing + inserted
    # Mirror the full catalog into every sub-team under this workspace so
    # sub-team role dropdowns show the unified list.
    for role in existing:
        await _mirror_catalog_role_to_subteams(
            workspace_id=team_id,
            role_name=role["name"],
            team_permissions=role.get("team_permissions") or {},
            is_system=bool(role.get("is_system")),
            is_default=bool(role.get("is_default")),
        )
    return existing


async def _mirror_catalog_role_to_subteams(
    workspace_id: str,
    role_name: str,
    team_permissions: dict,
    is_system: bool,
    is_default: bool,
) -> None:
    """For every sub-team under `workspace_id`, ensure a
    workspace_team_roles row exists with this name. If present, update its
    permissions / flags to match the catalog. If absent, insert it.
    Mirror rows share the same NAME as the catalog, but each sub-team has
    its own id (existing member rows reference those ids).
    """
    subteams = await _svc_select(
        "workspace_teams",
        {"workspace_id": f"eq.{workspace_id}", "select": "id"},
    )
    if not subteams:
        return
    payload = team_permissions or {}
    for st in subteams:
        existing = await _svc_select(
            "workspace_team_roles",
            {
                "team_id": f"eq.{st['id']}",
                "name": f"eq.{role_name}",
                "select": "id",
                "limit": "1",
            },
        )
        if existing:
            await _svc_update(
                "workspace_team_roles",
                {"id": f"eq.{existing[0]['id']}"},
                {
                    "permissions": payload,
                    "is_system": is_system,
                    "is_default": is_default,
                },
            )
        else:
            await _svc_insert(
                "workspace_team_roles",
                {
                    "team_id": st["id"],
                    "name": role_name,
                    "permissions": payload,
                    "is_system": is_system,
                    "is_default": is_default,
                },
            )


async def _unmirror_catalog_role(workspace_id: str, role_name: str) -> None:
    """Delete the mirror row for `role_name` in every sub-team under
    `workspace_id`. Before deleting, reassign any sub-team members on that
    role to the sub-team's default role so we don't strand them.
    """
    subteams = await _svc_select(
        "workspace_teams",
        {"workspace_id": f"eq.{workspace_id}", "select": "id"},
    )
    for st in subteams:
        mirror = await _svc_select(
            "workspace_team_roles",
            {
                "team_id": f"eq.{st['id']}",
                "name": f"eq.{role_name}",
                "select": "id",
                "limit": "1",
            },
        )
        if not mirror:
            continue
        role_id = mirror[0]["id"]
        # Find a fallback role in this sub-team (default → Admin → any other)
        others = await _svc_select(
            "workspace_team_roles",
            {"team_id": f"eq.{st['id']}", "select": "*"},
        )
        fallback = next(
            (r for r in others if r.get("is_default") and r["id"] != role_id),
            next(
                (r for r in others if r["name"] == "Admin" and r["id"] != role_id),
                next((r for r in others if r["id"] != role_id), None),
            ),
        )
        if fallback:
            await _svc_update(
                "workspace_team_members",
                {
                    "team_id": f"eq.{st['id']}",
                    "role_id": f"eq.{role_id}",
                },
                {"role_id": fallback["id"]},
            )
        await _svc_delete(
            "workspace_team_roles",
            {"id": f"eq.{role_id}", "team_id": f"eq.{st['id']}"},
        )


# ---------- Team CRUD ----------


async def _seed_default_workspace(user: UserContext) -> None:
    """First-run bootstrap: create default workspace1 + team1 with Admin/Role 1/2/3
    roles at both levels, and assign the user as Admin. Idempotent — caller must
    only invoke when the user has no owned workspaces and no memberships.
    """
    # 1. Create workspace1 (slug suffixed to avoid UNIQUE collision across users)
    slug = f"workspace1-{user.user_id[:8]}"
    ws_rows = await _svc_insert(
        "teams",
        {"name": "workspace1", "slug": slug, "owner_id": user.user_id},
    )
    ws = ws_rows[0]

    # 2. Seed workspace roles (Owner, Admin, Role 1/2/3) + add user as Owner
    ws_roles = await _ensure_default_roles(ws["id"])
    ws_owner = next((r for r in ws_roles if r["name"] == "Owner"), ws_roles[0])
    await _svc_insert(
        "team_members",
        {
            "team_id": ws["id"],
            "user_id": user.user_id,
            "role_id": ws_owner["id"],
            "added_by": user.user_id,
        },
    )

    # 3. Create team1 sub-team inside workspace1
    st_rows = await _svc_insert(
        "workspace_teams",
        {
            "workspace_id": ws["id"],
            "name": "team1",
            "description": "",
            "created_by": user.user_id,
        },
    )
    st = st_rows[0]

    # 4. Seed sub-team roles (Team Admin, Admin, Role 1/2/3) + add user as Team Admin
    st_roles = await _ensure_default_subteam_roles(st["id"])
    st_admin = next(
        (r for r in st_roles if r["name"] == "Team Admin"), st_roles[0]
    )
    await _svc_insert(
        "workspace_team_members",
        {
            "team_id": st["id"],
            "user_id": user.user_id,
            "role_id": st_admin["id"],
            "added_by": user.user_id,
        },
    )


@router.get("/teams")
async def v2_teams_list(user: UserContext = Depends(get_current_user)):
    """All teams the user owns or is a member of."""
    _require_service_role()
    owned = await _svc_select(
        "teams", {"owner_id": f"eq.{user.user_id}", "select": "*"}
    )
    memberships = await _svc_select(
        "team_members",
        {"user_id": f"eq.{user.user_id}", "select": "team_id,role_id,team_roles(name,permissions)"},
    )
    # First-run bootstrap: if the user has no workspaces at all, seed the
    # default workspace1 + team1 with Admin/Role 1/2/3 at both levels.
    if not owned and not memberships:
        try:
            await _seed_default_workspace(user)
            owned = await _svc_select(
                "teams", {"owner_id": f"eq.{user.user_id}", "select": "*"}
            )
            memberships = await _svc_select(
                "team_members",
                {"user_id": f"eq.{user.user_id}", "select": "team_id,role_id,team_roles(name,permissions)"},
            )
        except Exception as _seed_err:  # noqa: BLE001
            # Never let seeding failure block the dashboard — user can create
            # a workspace manually from the UI.
            print(f"[v2] seed_default_workspace failed for {user.user_id}: {_seed_err}")
    member_team_ids = [m["team_id"] for m in memberships]
    member_teams: list[dict] = []
    if member_team_ids:
        member_teams = await _svc_select(
            "teams",
            {"id": f"in.({','.join(member_team_ids)})", "select": "*"},
        )
    seen: set[str] = set()
    out: list[dict] = []
    for t in owned + member_teams:
        if t["id"] in seen:
            continue
        seen.add(t["id"])
        mem = next((m for m in memberships if m["team_id"] == t["id"]), None)
        role = (mem or {}).get("team_roles") or {}
        t["role"] = role.get("name") or ("Owner" if t["owner_id"] == user.user_id else "Member")
        t["is_owner"] = t["owner_id"] == user.user_id
        t["permissions"] = role.get("permissions") or (
            DEFAULT_ROLES[0]["permissions"] if t["is_owner"] else {}
        )
        out.append(t)
    out.sort(key=lambda t: t.get("created_at") or "", reverse=True)
    return {"teams": out}


@router.post("/teams")
async def v2_teams_create(
    payload: dict, user: UserContext = Depends(get_current_user)
):
    """Create a new team. Caller becomes owner with full permissions."""
    _require_service_role()
    name = (payload or {}).get("name", "").strip()
    if not name:
        raise HTTPException(400, "name required")
    slug = (payload or {}).get("slug") or name.lower().replace(" ", "-")[:40]
    rows = await _svc_insert(
        "teams", {"name": name, "slug": slug, "owner_id": user.user_id}
    )
    team = rows[0]
    roles = await _ensure_default_roles(team["id"])
    owner_role = next((r for r in roles if r["name"] == "Owner"), roles[0])
    await _svc_insert(
        "team_members",
        {
            "team_id": team["id"],
            "user_id": user.user_id,
            "role_id": owner_role["id"],
            "added_by": user.user_id,
        },
    )
    team["role"] = "Owner"
    team["is_owner"] = True
    team["permissions"] = owner_role["permissions"]
    return {"team": team}


@router.get("/teams/{team_id}")
async def v2_teams_get(team_id: str, user: UserContext = Depends(get_current_user)):
    team, m = await _require_member(team_id, user)
    role = (m or {}).get("team_roles") or {}
    team["role"] = role.get("name") or ("Owner" if team["owner_id"] == user.user_id else "Member")
    team["is_owner"] = team["owner_id"] == user.user_id
    team["permissions"] = role.get("permissions") or (
        DEFAULT_ROLES[0]["permissions"] if team["is_owner"] else {}
    )
    return {"team": team}


@router.patch("/teams/{team_id}")
async def v2_teams_update(
    team_id: str, payload: dict, user: UserContext = Depends(get_current_user)
):
    await _require_perm(team_id, user, "admin")
    updates: dict[str, Any] = {}
    if "name" in payload and str(payload["name"]).strip():
        updates["name"] = str(payload["name"]).strip()
    if "slug" in payload and str(payload["slug"]).strip():
        updates["slug"] = str(payload["slug"]).strip()
    if not updates:
        raise HTTPException(400, "Nothing to update")
    rows = await _svc_update(
        "teams", {"id": f"eq.{team_id}"}, updates
    )
    return {"team": rows[0] if rows else None}


@router.delete("/teams/{team_id}")
async def v2_teams_delete(team_id: str, user: UserContext = Depends(get_current_user)):
    """Owner-only. Cascades to roles/members/invitations."""
    team = await _get_team_or_404(team_id)
    if team["owner_id"] != user.user_id:
        raise HTTPException(403, "Only the owner can delete a team")
    await _svc_delete("teams", {"id": f"eq.{team_id}"})
    return {"ok": True}


# ---------- Roles ----------

@router.get("/teams/{team_id}/roles")
async def v2_roles_list(team_id: str, user: UserContext = Depends(get_current_user)):
    await _require_member(team_id, user)
    roles = await _ensure_default_roles(team_id)
    roles.sort(key=lambda r: (not r.get("is_system"), r.get("name", "")))
    return {"roles": roles}


@router.post("/teams/{team_id}/roles")
async def v2_roles_create(
    team_id: str, payload: dict, user: UserContext = Depends(get_current_user)
):
    await _require_perm(team_id, user, "manage_roles")
    name = (payload or {}).get("name", "").strip()
    if not name:
        raise HTTPException(400, "Role name required")
    perms = (payload or {}).get("permissions") or {}
    if not isinstance(perms, dict):
        raise HTTPException(400, "permissions must be an object")
    team_perms = (payload or {}).get("team_permissions") or {}
    if not isinstance(team_perms, dict):
        raise HTTPException(400, "team_permissions must be an object")
    rows = await _svc_insert(
        "team_roles",
        {
            "team_id": team_id,
            "name": name,
            "permissions": perms,
            "team_permissions": team_perms,
            "is_default": False,
            "is_system": False,
        },
    )
    # Mirror the new catalog role into every sub-team under this workspace.
    await _mirror_catalog_role_to_subteams(
        workspace_id=team_id,
        role_name=name,
        team_permissions=team_perms,
        is_system=False,
        is_default=False,
    )
    return {"role": rows[0] if rows else None}


@router.patch("/teams/{team_id}/roles/{role_id}")
async def v2_roles_update(
    team_id: str,
    role_id: str,
    payload: dict,
    user: UserContext = Depends(get_current_user),
):
    await _require_perm(team_id, user, "manage_roles")
    existing = await _svc_select(
        "team_roles",
        {"id": f"eq.{role_id}", "team_id": f"eq.{team_id}", "select": "*", "limit": "1"},
    )
    if not existing:
        raise HTTPException(404, "Role not found in this team")
    role = existing[0]
    updates: dict[str, Any] = {}
    if role.get("is_system") and role.get("name") == "Owner":
        raise HTTPException(400, "Owner role cannot be edited")
    if "name" in payload and str(payload["name"]).strip():
        if role.get("is_system"):
            raise HTTPException(400, "System role names cannot be renamed")
        updates["name"] = str(payload["name"]).strip()
    if "permissions" in payload and isinstance(payload["permissions"], dict):
        updates["permissions"] = payload["permissions"]
    if "team_permissions" in payload and isinstance(
        payload["team_permissions"], dict
    ):
        updates["team_permissions"] = payload["team_permissions"]
    if "is_default" in payload:
        updates["is_default"] = bool(payload["is_default"])
    if not updates:
        raise HTTPException(400, "Nothing to update")
    rows = await _svc_update(
        "team_roles",
        {"id": f"eq.{role_id}", "team_id": f"eq.{team_id}"},
        updates,
    )
    updated = rows[0] if rows else role
    # If the role was renamed, remove mirror rows under the OLD name first
    # so they don't orphan. The mirror helper below then re-creates rows
    # under the NEW name in every sub-team.
    old_name = role.get("name")
    new_name = updated.get("name") or old_name
    if old_name and old_name != new_name:
        await _unmirror_catalog_role(team_id, old_name)
    # Mirror the (possibly-renamed / repermed) catalog row into every
    # sub-team so mirror perms / flags stay in sync with the catalog.
    await _mirror_catalog_role_to_subteams(
        workspace_id=team_id,
        role_name=new_name,
        team_permissions=updated.get("team_permissions") or {},
        is_system=bool(updated.get("is_system")),
        is_default=bool(updated.get("is_default")),
    )
    return {"role": updated}


@router.delete("/teams/{team_id}/roles/{role_id}")
async def v2_roles_delete(
    team_id: str,
    role_id: str,
    user: UserContext = Depends(get_current_user),
):
    await _require_perm(team_id, user, "manage_roles")
    existing = await _svc_select(
        "team_roles",
        {"id": f"eq.{role_id}", "team_id": f"eq.{team_id}", "select": "*", "limit": "1"},
    )
    if not existing:
        raise HTTPException(404, "Role not found in this team")
    role = existing[0]
    if role.get("is_system"):
        raise HTTPException(400, "System roles (Owner/Admin/Member/Viewer) cannot be deleted")
    # Any members on this role get reassigned to the default role.
    all_roles = await _svc_select(
        "team_roles", {"team_id": f"eq.{team_id}", "select": "*"}
    )
    default = next(
        (r for r in all_roles if r.get("is_default") and r["id"] != role_id),
        next(
            (r for r in all_roles if r["name"] == "Admin" and r["id"] != role_id),
            next((r for r in all_roles if r["id"] != role_id), None),
        ),
    )
    if default:
        await _svc_update(
            "team_members",
            {"team_id": f"eq.{team_id}", "role_id": f"eq.{role_id}"},
            {"role_id": default["id"]},
        )
    await _svc_delete(
        "team_roles", {"id": f"eq.{role_id}", "team_id": f"eq.{team_id}"}
    )
    # Remove the mirror of this role from every sub-team and reassign any
    # sub-team members on it to the sub-team's fallback role.
    await _unmirror_catalog_role(team_id, role.get("name") or "")
    return {"ok": True}


# ---------- Members ----------

@router.get("/teams/{team_id}/members")
async def v2_members_list(
    team_id: str, user: UserContext = Depends(get_current_user)
):
    await _require_member(team_id, user)
    rows = await _svc_select(
        "team_members",
        {
            "team_id": f"eq.{team_id}",
            "select": "*,team_roles(id,name,permissions,is_system)",
            "order": "added_at.asc",
        },
    )
    team = await _get_team_or_404(team_id)
    for r in rows:
        r["is_owner"] = r["user_id"] == team["owner_id"]
    await _enrich_members_with_identity(rows)
    return {"members": rows, "owner_id": team["owner_id"]}


@router.patch("/teams/{team_id}/members/{member_id}")
async def v2_members_update(
    team_id: str,
    member_id: str,
    payload: dict,
    user: UserContext = Depends(get_current_user),
):
    await _require_perm(team_id, user, "manage_members")
    existing = await _svc_select(
        "team_members",
        {"id": f"eq.{member_id}", "team_id": f"eq.{team_id}", "select": "*", "limit": "1"},
    )
    if not existing:
        raise HTTPException(404, "Member not found")
    team = await _get_team_or_404(team_id)
    if existing[0]["user_id"] == team["owner_id"]:
        raise HTTPException(400, "Owner's role cannot be changed")
    role_id = (payload or {}).get("role_id")
    if not role_id:
        raise HTTPException(400, "role_id required")
    role_chk = await _svc_select(
        "team_roles",
        {"id": f"eq.{role_id}", "team_id": f"eq.{team_id}", "select": "id,name", "limit": "1"},
    )
    if not role_chk:
        raise HTTPException(400, "role_id does not belong to this team")
    rows = await _svc_update(
        "team_members",
        {"id": f"eq.{member_id}", "team_id": f"eq.{team_id}"},
        {"role_id": role_id},
    )
    return {"member": rows[0] if rows else None}


@router.delete("/teams/{team_id}/members/{member_id}")
async def v2_members_delete(
    team_id: str,
    member_id: str,
    user: UserContext = Depends(get_current_user),
):
    existing = await _svc_select(
        "team_members",
        {"id": f"eq.{member_id}", "team_id": f"eq.{team_id}", "select": "*", "limit": "1"},
    )
    if not existing:
        raise HTTPException(404, "Member not found")
    team = await _get_team_or_404(team_id)
    target_user_id = existing[0]["user_id"]
    if target_user_id == team["owner_id"]:
        raise HTTPException(400, "Owner cannot be removed — transfer ownership or delete the team")
    # Self-leave is allowed without manage_members permission.
    if target_user_id != user.user_id:
        await _require_perm(team_id, user, "manage_members")
    await _svc_delete(
        "team_members", {"id": f"eq.{member_id}", "team_id": f"eq.{team_id}"}
    )
    return {"ok": True}


# ---------- Invitations ----------

def _invite_url(request_host: str | None, token: str) -> str:
    # Frontend opens /?invite=TOKEN and shows the accept-invite modal.
    base = os.environ.get("PUBLIC_BASE_URL", "").strip()
    if not base:
        base = request_host or "https://brokerflow-eyt8.onrender.com"
    base = base.rstrip("/")
    return f"{base}/?invite={token}"


@router.get("/teams/{team_id}/invitations")
async def v2_invites_list(
    team_id: str, user: UserContext = Depends(get_current_user)
):
    await _require_perm(team_id, user, "manage_members")
    rows = await _svc_select(
        "team_invitations",
        {
            "team_id": f"eq.{team_id}",
            "select": "*,team_roles(name,permissions)",
            "order": "created_at.desc",
        },
    )
    for r in rows:
        r["status"] = (
            "accepted" if r.get("accepted_at")
            else "revoked" if r.get("revoked_at")
            else "expired" if r.get("expires_at") and r["expires_at"] < datetime.now(timezone.utc).isoformat()
            else "pending"
        )
        r["invite_url"] = _invite_url(None, r["token"])
    return {"invitations": rows}


@router.post("/teams/{team_id}/invitations")
async def v2_invites_create(
    team_id: str,
    payload: dict,
    user: UserContext = Depends(get_current_user),
):
    await _require_perm(team_id, user, "manage_members")
    email = (payload or {}).get("email", "").strip().lower()
    role_id = (payload or {}).get("role_id")
    if not email or "@" not in email:
        raise HTTPException(400, "valid email required")
    if not role_id:
        # Fall back to the team's default role (Admin by default).
        roles = await _ensure_default_roles(team_id)
        default = (
            next((r for r in roles if r.get("is_default")), None)
            or next((r for r in roles if r["name"] == "Admin"), None)
            or roles[0]
        )
        role_id = default["id"]
    else:
        role_chk = await _svc_select(
            "team_roles",
            {"id": f"eq.{role_id}", "team_id": f"eq.{team_id}", "select": "id", "limit": "1"},
        )
        if not role_chk:
            raise HTTPException(400, "role_id does not belong to this team")
    token = secrets.token_urlsafe(24)
    expires = (datetime.now(timezone.utc) + timedelta(days=14)).isoformat()
    rows = await _svc_insert(
        "team_invitations",
        {
            "team_id": team_id,
            "email": email,
            "role_id": role_id,
            "token": token,
            "invited_by": user.user_id,
            "expires_at": expires,
        },
    )
    inv = rows[0]
    inv["invite_url"] = _invite_url(None, token)
    return {"invitation": inv}


@router.delete("/teams/{team_id}/invitations/{invite_id}")
async def v2_invites_revoke(
    team_id: str,
    invite_id: str,
    user: UserContext = Depends(get_current_user),
):
    await _require_perm(team_id, user, "manage_members")
    await _svc_update(
        "team_invitations",
        {"id": f"eq.{invite_id}", "team_id": f"eq.{team_id}"},
        {"revoked_at": datetime.now(timezone.utc).isoformat()},
    )
    return {"ok": True}


@router.get("/invitations/{token}")
async def v2_invite_preview(
    token: str, user: UserContext = Depends(get_current_user)
):
    """Show invite details (team name, role) so the user can decide to accept."""
    rows = await _svc_select(
        "team_invitations",
        {"token": f"eq.{token}", "select": "*,team_roles(name,permissions)", "limit": "1"},
    )
    if not rows:
        raise HTTPException(404, "Invitation not found")
    inv = rows[0]
    if inv.get("accepted_at"):
        raise HTTPException(410, "Invitation already accepted")
    if inv.get("revoked_at"):
        raise HTTPException(410, "Invitation was revoked")
    if inv.get("expires_at") and inv["expires_at"] < datetime.now(timezone.utc).isoformat():
        raise HTTPException(410, "Invitation expired")
    team = await _get_team_or_404(inv["team_id"])
    inv["team_name"] = team["name"]
    inv["email_matches"] = (user.email or "").lower() == (inv.get("email") or "").lower()
    return {"invitation": inv}


@router.post("/invitations/{token}/accept")
async def v2_invite_accept(
    token: str, user: UserContext = Depends(get_current_user)
):
    rows = await _svc_select(
        "team_invitations", {"token": f"eq.{token}", "select": "*", "limit": "1"}
    )
    if not rows:
        raise HTTPException(404, "Invitation not found")
    inv = rows[0]
    if inv.get("accepted_at"):
        raise HTTPException(410, "Invitation already accepted")
    if inv.get("revoked_at"):
        raise HTTPException(410, "Invitation was revoked")
    if inv.get("expires_at") and inv["expires_at"] < datetime.now(timezone.utc).isoformat():
        raise HTTPException(410, "Invitation expired")
    if (user.email or "").lower() != (inv.get("email") or "").lower():
        raise HTTPException(
            403,
            f"This invitation was sent to {inv.get('email')}. Sign in with that email to accept.",
        )
    # Upsert membership.
    existing = await _get_membership(inv["team_id"], user.user_id)
    if existing:
        await _svc_update(
            "team_members",
            {"id": f"eq.{existing['id']}"},
            {"role_id": inv.get("role_id")},
        )
    else:
        await _svc_insert(
            "team_members",
            {
                "team_id": inv["team_id"],
                "user_id": user.user_id,
                "role_id": inv.get("role_id"),
                "added_by": inv.get("invited_by"),
            },
        )
    await _svc_update(
        "team_invitations",
        {"id": f"eq.{inv['id']}"},
        {
            "accepted_at": datetime.now(timezone.utc).isoformat(),
            "accepted_by": user.user_id,
        },
    )
    team = await _get_team_or_404(inv["team_id"])
    return {"ok": True, "team": team}


# =================================================================
# WORKSPACES — alias layer. The existing `teams` table and its four
# endpoints (`/teams/*`) act as the top-level WORKSPACE. These aliases
# let the frontend speak in workspace language while reusing the same
# storage, RLS, and permission helpers.
# =================================================================

@router.get("/workspaces")
async def v2_workspaces_list(user: UserContext = Depends(get_current_user)):
    return await v2_teams_list(user=user)


@router.post("/workspaces")
async def v2_workspaces_create(
    payload: dict, user: UserContext = Depends(get_current_user)
):
    return await v2_teams_create(payload=payload, user=user)


@router.get("/workspaces/{workspace_id}")
async def v2_workspaces_get(
    workspace_id: str, user: UserContext = Depends(get_current_user)
):
    return await v2_teams_get(team_id=workspace_id, user=user)


@router.patch("/workspaces/{workspace_id}")
async def v2_workspaces_update(
    workspace_id: str, payload: dict, user: UserContext = Depends(get_current_user)
):
    return await v2_teams_update(team_id=workspace_id, payload=payload, user=user)


@router.delete("/workspaces/{workspace_id}")
async def v2_workspaces_delete(
    workspace_id: str, user: UserContext = Depends(get_current_user)
):
    return await v2_teams_delete(team_id=workspace_id, user=user)


@router.get("/workspaces/{workspace_id}/members")
async def v2_workspace_members_list(
    workspace_id: str, user: UserContext = Depends(get_current_user)
):
    return await v2_members_list(team_id=workspace_id, user=user)


@router.patch("/workspaces/{workspace_id}/members/{member_id}")
async def v2_workspace_members_update(
    workspace_id: str,
    member_id: str,
    payload: dict,
    user: UserContext = Depends(get_current_user),
):
    return await v2_members_update(
        team_id=workspace_id, member_id=member_id, payload=payload, user=user
    )


@router.delete("/workspaces/{workspace_id}/members/{member_id}")
async def v2_workspace_members_delete(
    workspace_id: str,
    member_id: str,
    user: UserContext = Depends(get_current_user),
):
    return await v2_members_delete(
        team_id=workspace_id, member_id=member_id, user=user
    )


@router.get("/workspaces/{workspace_id}/roles")
async def v2_workspace_roles_list(
    workspace_id: str, user: UserContext = Depends(get_current_user)
):
    return await v2_roles_list(team_id=workspace_id, user=user)


@router.post("/workspaces/{workspace_id}/roles")
async def v2_workspace_roles_create(
    workspace_id: str, payload: dict, user: UserContext = Depends(get_current_user)
):
    return await v2_roles_create(team_id=workspace_id, payload=payload, user=user)


@router.patch("/workspaces/{workspace_id}/roles/{role_id}")
async def v2_workspace_roles_update(
    workspace_id: str,
    role_id: str,
    payload: dict,
    user: UserContext = Depends(get_current_user),
):
    return await v2_roles_update(
        team_id=workspace_id, role_id=role_id, payload=payload, user=user
    )


@router.delete("/workspaces/{workspace_id}/roles/{role_id}")
async def v2_workspace_roles_delete(
    workspace_id: str,
    role_id: str,
    user: UserContext = Depends(get_current_user),
):
    return await v2_roles_delete(team_id=workspace_id, role_id=role_id, user=user)


@router.get("/workspaces/{workspace_id}/invitations")
async def v2_workspace_invites_list(
    workspace_id: str, user: UserContext = Depends(get_current_user)
):
    return await v2_invites_list(team_id=workspace_id, user=user)


@router.post("/workspaces/{workspace_id}/invitations")
async def v2_workspace_invites_create(
    workspace_id: str, payload: dict, user: UserContext = Depends(get_current_user)
):
    return await v2_invites_create(team_id=workspace_id, payload=payload, user=user)


@router.delete("/workspaces/{workspace_id}/invitations/{invite_id}")
async def v2_workspace_invites_revoke(
    workspace_id: str,
    invite_id: str,
    user: UserContext = Depends(get_current_user),
):
    return await v2_invites_revoke(team_id=workspace_id, invite_id=invite_id, user=user)


@router.get("/workspaces/{workspace_id}/people")
async def v2_workspace_people(
    workspace_id: str, user: UserContext = Depends(get_current_user)
):
    """Flat "All people" view: one row per workspace member with their
    workspace role + every sub-team they belong to. Used by the Discord-style
    Settings → Team tab's default pane.
    """
    await _require_member(workspace_id, user)
    workspace = await _get_team_or_404(workspace_id)

    ws_members = await _svc_select(
        "team_members",
        {
            "team_id": f"eq.{workspace_id}",
            "select": "id,user_id,role_id,added_at,team_roles(id,name,permissions,is_system)",
            "order": "added_at.asc",
        },
    )
    subteams = await _svc_select(
        "workspace_teams",
        {
            "workspace_id": f"eq.{workspace_id}",
            "select": "id,name",
            "order": "created_at.asc",
        },
    )
    st_ids = [s["id"] for s in subteams]
    st_by_id = {s["id"]: s for s in subteams}
    subteam_memberships: list[dict] = []
    if st_ids:
        subteam_memberships = await _svc_select(
            "workspace_team_members",
            {
                "team_id": f"in.({','.join(st_ids)})",
                "select": "id,team_id,user_id,role_id,workspace_team_roles(id,name,is_system)",
            },
        )

    # Pending invites — show as "invited" rows so admins can resend/revoke
    invites = await _svc_select(
        "team_invitations",
        {
            "team_id": f"eq.{workspace_id}",
            "select": "id,email,role_id,token,expires_at,accepted_at,revoked_at,team_roles(name)",
            "order": "created_at.desc",
        },
    )

    # identity map: every workspace user_id
    ids = [m["user_id"] for m in ws_members]
    idmap = await _identity_map(ids)

    # group sub-team memberships by user_id
    by_user: dict[str, list[dict]] = {}
    for sm in subteam_memberships:
        uid = sm.get("user_id")
        if not uid:
            continue
        st = st_by_id.get(sm["team_id"]) or {}
        role = sm.get("workspace_team_roles") or {}
        by_user.setdefault(uid, []).append({
            "membership_id": sm["id"],
            "team_id": sm["team_id"],
            "team_name": st.get("name"),
            "role_id": sm.get("role_id"),
            "role_name": role.get("name"),
        })

    people = []
    for m in ws_members:
        uid = m["user_id"]
        ident = idmap.get(uid) or _identity_from_auth_user(None)
        role = m.get("team_roles") or {}
        people.append({
            "membership_id": m["id"],
            "user_id": uid,
            "email": ident["email"],
            "display_name": ident["display_name"],
            "avatar_url": ident["avatar_url"],
            "is_owner": uid == workspace["owner_id"],
            "ws_role_id": m.get("role_id"),
            "ws_role_name": role.get("name"),
            "ws_role_permissions": role.get("permissions") or {},
            "teams": by_user.get(uid, []),
        })

    pending = []
    now_iso = datetime.now(timezone.utc).isoformat()
    for inv in invites:
        if inv.get("accepted_at") or inv.get("revoked_at"):
            continue
        status = "expired" if inv.get("expires_at") and inv["expires_at"] < now_iso else "pending"
        role = inv.get("team_roles") or {}
        pending.append({
            "invitation_id": inv["id"],
            "email": inv["email"],
            "role_id": inv.get("role_id"),
            "role_name": role.get("name"),
            "status": status,
            "expires_at": inv.get("expires_at"),
            "token": inv["token"],
        })

    return {
        "workspace": {"id": workspace["id"], "name": workspace.get("name"), "owner_id": workspace["owner_id"]},
        "subteams": subteams,
        "people": people,
        "pending_invitations": pending,
    }


# =================================================================
# SUB-TEAMS inside a workspace — groupings like "Acquisitions",
# "Asset Mgmt", "Dispositions". Each has its own members and roles.
# A user must already be a workspace member to be added to a sub-team.
# Workspace admins/owners have full rights on every sub-team in their
# workspace; otherwise sub-team permissions are enforced by role.
# =================================================================

DEFAULT_SUBTEAM_ROLES = [
    {
        "name": "Team Admin",
        "is_system": True,
        "is_default": False,
        "permissions": {
            "team_admin": True,
            "manage_team_members": True,
            "manage_team_roles": True,
            "edit_team_content": True,
            "view_team_content": True,
        },
    },
    {
        "name": "Admin",
        "is_system": False,
        "is_default": True,
        "permissions": {
            "team_admin": True,
            "manage_team_members": True,
            "manage_team_roles": True,
            "edit_team_content": True,
            "view_team_content": True,
        },
    },
    {
        "name": "Role 1",
        "is_system": False,
        "is_default": False,
        "permissions": {
            "team_admin": False,
            "manage_team_members": False,
            "manage_team_roles": False,
            "edit_team_content": False,
            "view_team_content": False,
        },
    },
    {
        "name": "Role 2",
        "is_system": False,
        "is_default": False,
        "permissions": {
            "team_admin": False,
            "manage_team_members": False,
            "manage_team_roles": False,
            "edit_team_content": False,
            "view_team_content": False,
        },
    },
    {
        "name": "Role 3",
        "is_system": False,
        "is_default": False,
        "permissions": {
            "team_admin": False,
            "manage_team_members": False,
            "manage_team_roles": False,
            "edit_team_content": False,
            "view_team_content": False,
        },
    },
]


async def _get_subteam_or_404(team_id: str) -> dict:
    rows = await _svc_select(
        "workspace_teams", {"id": f"eq.{team_id}", "select": "*", "limit": "1"}
    )
    if not rows:
        raise HTTPException(404, "Team not found")
    return rows[0]


async def _get_subteam_membership(team_id: str, user_id: str) -> dict | None:
    rows = await _svc_select(
        "workspace_team_members",
        {
            "team_id": f"eq.{team_id}",
            "user_id": f"eq.{user_id}",
            "select": "*,workspace_team_roles(*)",
            "limit": "1",
        },
    )
    return rows[0] if rows else None


async def _ensure_default_subteam_roles(team_id: str) -> list[dict]:
    """Seed a new sub-team's role mirror from its parent workspace's
    unified catalog (public.team_roles). Each catalog row yields one
    mirror row in public.workspace_team_roles whose `permissions` column
    holds the catalog row's `team_permissions` (sub-team scope). Falls
    back to DEFAULT_SUBTEAM_ROLES only if the parent workspace has no
    catalog yet (shouldn't happen after first run, but safe).
    """
    subteam = await _svc_select(
        "workspace_teams",
        {"id": f"eq.{team_id}", "select": "workspace_id", "limit": "1"},
    )
    workspace_id = subteam[0]["workspace_id"] if subteam else None

    existing = await _svc_select(
        "workspace_team_roles", {"team_id": f"eq.{team_id}", "select": "*"}
    )
    existing_names = {r["name"] for r in existing}

    # Pull the parent workspace's unified catalog.
    catalog: list[dict] = []
    if workspace_id:
        catalog = await _svc_select(
            "team_roles", {"team_id": f"eq.{workspace_id}", "select": "*"}
        )

    if catalog:
        to_insert = [
            {
                "team_id": team_id,
                "name": r["name"],
                "permissions": r.get("team_permissions") or {},
                "is_default": bool(r.get("is_default")),
                "is_system": bool(r.get("is_system")),
            }
            for r in catalog
            if r["name"] not in existing_names
        ]
    else:
        # Fallback — workspace has no catalog yet (first-run edge case).
        to_insert = [
            {
                "team_id": team_id,
                "name": r["name"],
                "permissions": r["permissions"],
                "is_default": r["is_default"],
                "is_system": r["is_system"],
            }
            for r in DEFAULT_SUBTEAM_ROLES
            if r["name"] not in existing_names
        ]
    if to_insert:
        inserted = await _svc_insert("workspace_team_roles", to_insert)
        existing = existing + inserted
    return existing


async def _require_workspace_member_for_subteam(
    workspace_id: str, user: UserContext
) -> tuple[dict, dict]:
    """Verify user is a member (or owner) of the workspace. Returns
    (workspace, membership)."""
    return await _require_member(workspace_id, user)


async def _require_subteam_access(
    workspace_id: str, team_id: str, user: UserContext, perm: str | None = None
) -> tuple[dict, dict, dict, dict | None]:
    """Returns (workspace, workspace_membership, subteam, subteam_membership).
    Workspace owner and admins bypass subteam-level perm checks. Otherwise
    the user must be a subteam member with the requested permission.
    """
    workspace, ws_membership = await _require_workspace_member_for_subteam(
        workspace_id, user
    )
    subteam = await _get_subteam_or_404(team_id)
    if subteam["workspace_id"] != workspace_id:
        raise HTTPException(404, "Team not found in this workspace")

    # Workspace owner / admin has full rights on every sub-team.
    if workspace["owner_id"] == user.user_id or _has_perm(
        workspace, ws_membership, user, "admin"
    ):
        sub_m = await _get_subteam_membership(team_id, user.user_id)
        return workspace, ws_membership, subteam, sub_m

    sub_m = await _get_subteam_membership(team_id, user.user_id)
    if not sub_m:
        raise HTTPException(403, "Not a member of this team")

    if perm:
        role = (sub_m or {}).get("workspace_team_roles") or {}
        perms = role.get("permissions") or {}
        if not (perms.get("team_admin") or perms.get(perm)):
            raise HTTPException(403, f"Missing team permission: {perm}")

    return workspace, ws_membership, subteam, sub_m


# ---------- Sub-team CRUD ----------

@router.get("/workspaces/{workspace_id}/teams")
async def v2_subteams_list(
    workspace_id: str, user: UserContext = Depends(get_current_user)
):
    """List all sub-teams in the workspace the user can see.
    Workspace members see every sub-team; the per-team membership
    is surfaced so the UI can show "Joined" badges.
    """
    _require_service_role()
    await _require_workspace_member_for_subteam(workspace_id, user)
    subteams = await _svc_select(
        "workspace_teams",
        {
            "workspace_id": f"eq.{workspace_id}",
            "select": "*",
            "order": "created_at.desc",
        },
    )
    # Fetch the caller's membership rows in bulk so we can annotate.
    my_mems = await _svc_select(
        "workspace_team_members",
        {
            "user_id": f"eq.{user.user_id}",
            "select": "team_id,role_id,workspace_team_roles(name,permissions)",
        },
    )
    my_by_team = {m["team_id"]: m for m in my_mems}
    # Count members per team.
    counts = await _svc_select(
        "workspace_team_members",
        {
            "team_id": f"in.({','.join([s['id'] for s in subteams]) or 'none'})",
            "select": "team_id",
        },
    )
    count_map: dict[str, int] = {}
    for c in counts:
        count_map[c["team_id"]] = count_map.get(c["team_id"], 0) + 1
    for s in subteams:
        mem = my_by_team.get(s["id"])
        s["member_count"] = count_map.get(s["id"], 0)
        s["joined"] = bool(mem)
        s["my_role"] = (
            (mem.get("workspace_team_roles") or {}).get("name") if mem else None
        )
        s["my_permissions"] = (
            (mem.get("workspace_team_roles") or {}).get("permissions") if mem else {}
        )
    return {"teams": subteams}


@router.post("/workspaces/{workspace_id}/teams")
async def v2_subteams_create(
    workspace_id: str, payload: dict, user: UserContext = Depends(get_current_user)
):
    """Create a sub-team. Requires workspace manage_members or admin.
    Creator becomes the first Team Admin.
    """
    _require_service_role()
    await _require_perm(workspace_id, user, "manage_members")
    name = (payload or {}).get("name", "").strip()
    if not name:
        raise HTTPException(400, "name required")
    description = (payload or {}).get("description", "")
    rows = await _svc_insert(
        "workspace_teams",
        {
            "workspace_id": workspace_id,
            "name": name,
            "description": description,
            "created_by": user.user_id,
        },
    )
    subteam = rows[0]
    roles = await _ensure_default_subteam_roles(subteam["id"])
    admin_role = next(
        (r for r in roles if r["name"] == "Team Admin"), roles[0]
    )
    await _svc_insert(
        "workspace_team_members",
        {
            "team_id": subteam["id"],
            "user_id": user.user_id,
            "role_id": admin_role["id"],
            "added_by": user.user_id,
        },
    )
    subteam["member_count"] = 1
    subteam["joined"] = True
    subteam["my_role"] = "Team Admin"
    subteam["my_permissions"] = admin_role["permissions"]
    return {"team": subteam}


@router.get("/workspaces/{workspace_id}/teams/{team_id}")
async def v2_subteams_get(
    workspace_id: str,
    team_id: str,
    user: UserContext = Depends(get_current_user),
):
    _, _, subteam, sub_m = await _require_subteam_access(
        workspace_id, team_id, user
    )
    role = (sub_m or {}).get("workspace_team_roles") or {}
    subteam["joined"] = bool(sub_m)
    subteam["my_role"] = role.get("name")
    subteam["my_permissions"] = role.get("permissions") or {}
    return {"team": subteam}


@router.patch("/workspaces/{workspace_id}/teams/{team_id}")
async def v2_subteams_update(
    workspace_id: str,
    team_id: str,
    payload: dict,
    user: UserContext = Depends(get_current_user),
):
    await _require_subteam_access(workspace_id, team_id, user, perm="team_admin")
    updates: dict[str, Any] = {}
    if "name" in payload and str(payload["name"]).strip():
        updates["name"] = str(payload["name"]).strip()
    if "description" in payload:
        updates["description"] = str(payload["description"] or "")
    if not updates:
        raise HTTPException(400, "Nothing to update")
    rows = await _svc_update(
        "workspace_teams",
        {"id": f"eq.{team_id}", "workspace_id": f"eq.{workspace_id}"},
        updates,
    )
    return {"team": rows[0] if rows else None}


@router.delete("/workspaces/{workspace_id}/teams/{team_id}")
async def v2_subteams_delete(
    workspace_id: str,
    team_id: str,
    user: UserContext = Depends(get_current_user),
):
    # Only workspace admins / owner can fully delete a sub-team.
    await _require_perm(workspace_id, user, "admin")
    await _get_subteam_or_404(team_id)
    await _svc_delete(
        "workspace_teams",
        {"id": f"eq.{team_id}", "workspace_id": f"eq.{workspace_id}"},
    )
    return {"ok": True}


# ---------- Sub-team roles ----------

@router.get("/workspaces/{workspace_id}/teams/{team_id}/roles")
async def v2_subteam_roles_list(
    workspace_id: str,
    team_id: str,
    user: UserContext = Depends(get_current_user),
):
    await _require_subteam_access(workspace_id, team_id, user)
    roles = await _ensure_default_subteam_roles(team_id)
    roles.sort(key=lambda r: (not r.get("is_system"), r.get("name", "")))
    return {"roles": roles}


@router.post("/workspaces/{workspace_id}/teams/{team_id}/roles")
async def v2_subteam_roles_create(
    workspace_id: str,
    team_id: str,
    payload: dict,
    user: UserContext = Depends(get_current_user),
):
    await _require_subteam_access(
        workspace_id, team_id, user, perm="manage_team_roles"
    )
    name = (payload or {}).get("name", "").strip()
    if not name:
        raise HTTPException(400, "Role name required")
    perms = (payload or {}).get("permissions") or {}
    if not isinstance(perms, dict):
        raise HTTPException(400, "permissions must be an object")
    rows = await _svc_insert(
        "workspace_team_roles",
        {
            "team_id": team_id,
            "name": name,
            "permissions": perms,
            "is_default": False,
            "is_system": False,
        },
    )
    return {"role": rows[0] if rows else None}


@router.patch("/workspaces/{workspace_id}/teams/{team_id}/roles/{role_id}")
async def v2_subteam_roles_update(
    workspace_id: str,
    team_id: str,
    role_id: str,
    payload: dict,
    user: UserContext = Depends(get_current_user),
):
    await _require_subteam_access(
        workspace_id, team_id, user, perm="manage_team_roles"
    )
    existing = await _svc_select(
        "workspace_team_roles",
        {
            "id": f"eq.{role_id}",
            "team_id": f"eq.{team_id}",
            "select": "*",
            "limit": "1",
        },
    )
    if not existing:
        raise HTTPException(404, "Role not found in this team")
    role = existing[0]
    updates: dict[str, Any] = {}
    if "name" in payload and str(payload["name"]).strip():
        if role.get("is_system"):
            raise HTTPException(400, "System role names cannot be renamed")
        updates["name"] = str(payload["name"]).strip()
    if "permissions" in payload and isinstance(payload["permissions"], dict):
        updates["permissions"] = payload["permissions"]
    if "is_default" in payload:
        updates["is_default"] = bool(payload["is_default"])
    if not updates:
        raise HTTPException(400, "Nothing to update")
    rows = await _svc_update(
        "workspace_team_roles",
        {"id": f"eq.{role_id}", "team_id": f"eq.{team_id}"},
        updates,
    )
    return {"role": rows[0] if rows else None}


@router.delete("/workspaces/{workspace_id}/teams/{team_id}/roles/{role_id}")
async def v2_subteam_roles_delete(
    workspace_id: str,
    team_id: str,
    role_id: str,
    user: UserContext = Depends(get_current_user),
):
    await _require_subteam_access(
        workspace_id, team_id, user, perm="manage_team_roles"
    )
    existing = await _svc_select(
        "workspace_team_roles",
        {
            "id": f"eq.{role_id}",
            "team_id": f"eq.{team_id}",
            "select": "*",
            "limit": "1",
        },
    )
    if not existing:
        raise HTTPException(404, "Role not found in this team")
    role = existing[0]
    if role.get("is_system"):
        raise HTTPException(400, "System roles cannot be deleted")
    all_roles = await _svc_select(
        "workspace_team_roles", {"team_id": f"eq.{team_id}", "select": "*"}
    )
    default = next(
        (r for r in all_roles if r.get("is_default") and r["id"] != role_id),
        next(
            (r for r in all_roles if r["name"] == "Admin" and r["id"] != role_id),
            next((r for r in all_roles if r["id"] != role_id), None),
        ),
    )
    if default:
        await _svc_update(
            "workspace_team_members",
            {"team_id": f"eq.{team_id}", "role_id": f"eq.{role_id}"},
            {"role_id": default["id"]},
        )
    await _svc_delete(
        "workspace_team_roles",
        {"id": f"eq.{role_id}", "team_id": f"eq.{team_id}"},
    )
    return {"ok": True}


# ---------- Sub-team members ----------

@router.get("/workspaces/{workspace_id}/teams/{team_id}/members")
async def v2_subteam_members_list(
    workspace_id: str,
    team_id: str,
    user: UserContext = Depends(get_current_user),
):
    await _require_subteam_access(workspace_id, team_id, user)
    rows = await _svc_select(
        "workspace_team_members",
        {
            "team_id": f"eq.{team_id}",
            "select": "*,workspace_team_roles(id,name,permissions,is_system)",
            "order": "added_at.asc",
        },
    )
    await _enrich_members_with_identity(rows)
    return {"members": rows}


@router.post("/workspaces/{workspace_id}/teams/{team_id}/members")
async def v2_subteam_members_add(
    workspace_id: str,
    team_id: str,
    payload: dict,
    user: UserContext = Depends(get_current_user),
):
    """Add an existing workspace member to a sub-team."""
    await _require_subteam_access(
        workspace_id, team_id, user, perm="manage_team_members"
    )
    target_user_id = (payload or {}).get("user_id")
    role_id = (payload or {}).get("role_id")
    if not target_user_id:
        raise HTTPException(400, "user_id required")

    # Target must already belong to the workspace.
    ws_membership = await _get_membership(workspace_id, target_user_id)
    workspace = await _get_team_or_404(workspace_id)
    if not ws_membership and workspace["owner_id"] != target_user_id:
        raise HTTPException(
            400,
            "User is not a member of this workspace. Invite them to the workspace first.",
        )

    if not role_id:
        roles = await _ensure_default_subteam_roles(team_id)
        default = next((r for r in roles if r.get("is_default")), roles[0])
        role_id = default["id"]
    else:
        chk = await _svc_select(
            "workspace_team_roles",
            {
                "id": f"eq.{role_id}",
                "team_id": f"eq.{team_id}",
                "select": "id",
                "limit": "1",
            },
        )
        if not chk:
            raise HTTPException(400, "role_id does not belong to this team")

    # Upsert — if already a member, update their role.
    existing = await _get_subteam_membership(team_id, target_user_id)
    if existing:
        rows = await _svc_update(
            "workspace_team_members",
            {"id": f"eq.{existing['id']}"},
            {"role_id": role_id},
        )
        return {"member": rows[0] if rows else None, "updated": True}
    rows = await _svc_insert(
        "workspace_team_members",
        {
            "team_id": team_id,
            "user_id": target_user_id,
            "role_id": role_id,
            "added_by": user.user_id,
        },
    )
    return {"member": rows[0] if rows else None, "updated": False}


@router.patch("/workspaces/{workspace_id}/teams/{team_id}/members/{member_id}")
async def v2_subteam_members_update(
    workspace_id: str,
    team_id: str,
    member_id: str,
    payload: dict,
    user: UserContext = Depends(get_current_user),
):
    await _require_subteam_access(
        workspace_id, team_id, user, perm="manage_team_members"
    )
    existing = await _svc_select(
        "workspace_team_members",
        {
            "id": f"eq.{member_id}",
            "team_id": f"eq.{team_id}",
            "select": "*",
            "limit": "1",
        },
    )
    if not existing:
        raise HTTPException(404, "Member not found")
    role_id = (payload or {}).get("role_id")
    if not role_id:
        raise HTTPException(400, "role_id required")
    chk = await _svc_select(
        "workspace_team_roles",
        {
            "id": f"eq.{role_id}",
            "team_id": f"eq.{team_id}",
            "select": "id",
            "limit": "1",
        },
    )
    if not chk:
        raise HTTPException(400, "role_id does not belong to this team")
    rows = await _svc_update(
        "workspace_team_members",
        {"id": f"eq.{member_id}", "team_id": f"eq.{team_id}"},
        {"role_id": role_id},
    )
    return {"member": rows[0] if rows else None}


@router.delete("/workspaces/{workspace_id}/teams/{team_id}/members/{member_id}")
async def v2_subteam_members_remove(
    workspace_id: str,
    team_id: str,
    member_id: str,
    user: UserContext = Depends(get_current_user),
):
    existing = await _svc_select(
        "workspace_team_members",
        {
            "id": f"eq.{member_id}",
            "team_id": f"eq.{team_id}",
            "select": "*",
            "limit": "1",
        },
    )
    if not existing:
        raise HTTPException(404, "Member not found")
    # Self-leave always allowed. Otherwise need manage_team_members.
    if existing[0]["user_id"] != user.user_id:
        await _require_subteam_access(
            workspace_id, team_id, user, perm="manage_team_members"
        )
    else:
        await _require_workspace_member_for_subteam(workspace_id, user)
    await _svc_delete(
        "workspace_team_members",
        {"id": f"eq.{member_id}", "team_id": f"eq.{team_id}"},
    )
    return {"ok": True}
